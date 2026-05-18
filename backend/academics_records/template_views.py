# academics_records/template_views.py

"""
Template generation endpoints for grade and attendance Excel files.

Grade template logic:
  1. Teacher submits: academic_year, term, school_level, class_level, subject, grade_type
  2. System looks up the teacher's TeacherAssignment for that subject/class/year/term
     and collects the classrooms the teacher is assigned to teach.
  3. Students are fetched who:
       - belong to that academic_year + term + school_level + class_level
       - are assigned to one of those classrooms (via StudentClassroomAssignment)
  4. The downloaded Excel has columns:
       Roll Number | Full Name | <grade_type label> (score)
     One row per student, sorted by full_name.

Endpoints:
  GET  /api/academics-records/templates/grades/
  GET  /api/academics-records/templates/attendance/
  GET  /api/academics-records/templates/trimesters/
"""

import io
import sys
import calendar
import traceback
from datetime import date, datetime

from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from academics.models import AcademicYear, ClassLevel, Subject, ClassRoom, Term, SchoolLevel
from students.models import Student, StudentClassroomAssignment
from teachers.models import Teacher, TeacherAssignment

from .models import GradeType
from .translations import get_lang, t


# =============================================================================
#  Error helper — logs to terminal AND returns a rich JSON response
# =============================================================================

def _error(message: str, http_status=status.HTTP_400_BAD_REQUEST,
           lang: str = "en", detail: str = None, exc: Exception = None) -> Response:
    """
    Print a structured error to stderr (in whatever language the request used),
    then return a DRF Response with full context.

    Terminal output:
        [ERROR 403] [lang=fr] Teacher not assigned to subject X in class Y
          [DETAIL]    No TeacherAssignment row exists for teacher_id=3, ...
          [EXCEPTION] DoesNotExist: ...
          [TRACEBACK] ...

    JSON response body:
        {
            "success": false,
            "message": "<localised message>",
            "detail":  "<extra context>",
            "status":  403
        }
    """
    payload = {
        "success": False,
        "message": message,
        "status":  http_status,
    }
    if detail:
        payload["detail"] = detail
    elif exc is not None:
        payload["detail"] = str(exc)

    # --- terminal -----------------------------------------------------------
    lines = [f"[ERROR {http_status}] [lang={lang}] {message}"]
    if detail:
        lines.append(f"  [DETAIL]    {detail}")
    if exc is not None:
        lines.append(f"  [EXCEPTION] {type(exc).__name__}: {exc}")
        tb = traceback.format_exc()
        if tb and tb.strip() not in ("NoneType: None", ""):
            lines.append(f"  [TRACEBACK]\n{tb}")
    print("\n".join(lines), file=sys.stderr, flush=True)
    # ------------------------------------------------------------------------

    return Response(payload, status=http_status)


# =============================================================================
#  Rwanda trimester definitions
# =============================================================================

RWANDA_TRIMESTERS = [
    {
        "code": "T1",
        "label_en": "Trimester 1 (January – April)",
        "label_fr": "Trimestre 1 (Janvier – Avril)",
        "label_rw": "Igice cya 1 (Mutarama – Mata)",
        "month_start": 1, "month_end": 4,
    },
    {
        "code": "T2",
        "label_en": "Trimester 2 (May – August)",
        "label_fr": "Trimestre 2 (Mai – Août)",
        "label_rw": "Igice cya 2 (Gicurasi – Kanama)",
        "month_start": 5, "month_end": 8,
    },
    {
        "code": "T3",
        "label_en": "Trimester 3 (September – November)",
        "label_fr": "Trimestre 3 (Septembre – Novembre)",
        "label_rw": "Igice cya 3 (Nzeri – Ugushyingo)",
        "month_start": 9, "month_end": 11,
    },
]


def _trimester_dates(academic_year: AcademicYear, trimester_code: str):
    tri = next((t for t in RWANDA_TRIMESTERS if t["code"] == trimester_code), None)
    if not tri:
        return academic_year.start_date, academic_year.end_date
    year     = academic_year.start_date.year
    start    = date(year, tri["month_start"], 1)
    last_day = calendar.monthrange(year, tri["month_end"])[1]
    end      = date(year, tri["month_end"], last_day)
    return max(start, academic_year.start_date), min(end, academic_year.end_date)


def _tri_label(tri: dict, lang: str) -> str:
    if lang == "fr": return tri.get("label_fr", tri["label_en"])
    if lang == "rw": return tri.get("label_rw", tri["label_en"])
    return tri["label_en"]


# =============================================================================
#  openpyxl style constants
# =============================================================================

_HDR_FILL   = PatternFill("solid", start_color="1F4E79")
_META_FILL  = PatternFill("solid", start_color="D6E4F0")
_COL_FILL   = PatternFill("solid", start_color="2E75B6")
_LOCK_FILL  = PatternFill("solid", start_color="F2F2F2")
_INPUT_FILL = PatternFill("solid", start_color="E2EFDA")
_WARN_FILL  = PatternFill("solid", start_color="FFF2CC")

_WHITE_BOLD = Font(name="Arial", bold=True,   color="FFFFFF", size=11)
_DARK_BOLD  = Font(name="Arial", bold=True,   color="1F4E79", size=10)
_DARK_REG   = Font(name="Arial",              color="404040", size=10)
_RED_BOLD   = Font(name="Arial", bold=True,   color="C00000", size=10)
_GREY_ITAL  = Font(name="Arial", italic=True, color="7F7F7F", size=9)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _cell(ws, row, col, value=None, font=None, fill=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = _THIN
    return c


def _merged_label_value(ws, row, label, value,
                         label_font, label_fill,
                         value_font, value_fill,
                         merge_start_col=2, merge_end_col=3):
    """
    Write label in column 1, then merge cols merge_start_col:merge_end_col
    BEFORE writing the value — prevents ValueError on newer openpyxl versions
    where merging would otherwise clear the cell value.
    """
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font = label_font; lbl.fill = label_fill
    lbl.alignment = _LEFT; lbl.border = _THIN

    ws.merge_cells(                         # ← merge FIRST
        start_row=row, start_column=merge_start_col,
        end_row=row,   end_column=merge_end_col,
    )
    val = ws.cell(row=row, column=merge_start_col,   # ← then write
                  value=str(value) if value is not None else "")
    val.font = value_font; val.fill = value_fill
    val.alignment = _LEFT; val.border = _THIN
    return val


# =============================================================================
#  Auth / teacher helpers
# =============================================================================

def _is_admin(user):
    return user.role == "admin" or user.is_superuser or user.is_staff


def _get_teacher(user, lang):
    """Return (Teacher, None) or (None, error Response)."""
    try:
        return Teacher.objects.get(user=user), None
    except Teacher.DoesNotExist as exc:
        return None, _error(
            t("teacher_profile_not_found", lang),
            status.HTTP_404_NOT_FOUND, lang=lang, exc=exc,
            detail=f"No Teacher row linked to user id={user.id} (username={user.username})",
        )


def _check_teacher_assignment(teacher, academic_year, class_level, subject,
                               lang, term=None):
    """
    Verify the teacher has an active assignment for subject + class.
    Prints a detailed diagnostic block to stderr on every call.
    Returns an error Response if not assigned, else None.
    """
    qs = TeacherAssignment.objects.filter(
        teacher=teacher,
        academic_year=academic_year,
        class_level=class_level,
        subject=subject,
    )

    # Always log the check so you can follow what happened in the terminal
    print(
        f"  [ASSIGNMENT CHECK]"
        f" teacher={teacher.full_name!r} (id={teacher.id})"
        f" | year={academic_year.name!r} (id={academic_year.id})"
        f" | class={class_level.name!r} (id={class_level.id})"
        f" | subject={subject.name!r} (id={subject.id})"
        f" | term={'none' if not term else f'{term.name!r} (id={term.id})'}",
        file=sys.stderr, flush=True,
    )
    all_rows = list(qs.values("id", "status", "term_id"))
    print(f"  [ASSIGNMENT ROWS] {all_rows or 'none'}", file=sys.stderr, flush=True)

    active_qs = qs.filter(status="active")
    if term:
        if active_qs.filter(term=term).exists():
            return None
        if active_qs.filter(term__isnull=True).exists():
            return None   # assignment has no term restriction → still valid
    else:
        if active_qs.exists():
            return None

    # Build an actionable explanation
    if not qs.exists():
        detail = (
            f"No TeacherAssignment row exists for "
            f"teacher_id={teacher.id}, academic_year_id={academic_year.id}, "
            f"class_level_id={class_level.id}, subject_id={subject.id}. "
            "Please create one in the timetable / admin panel."
        )
    else:
        statuses = list(qs.values_list("status", flat=True).distinct())
        terms    = list(qs.values_list("term__name", flat=True).distinct())
        detail = (
            f"Assignment row(s) exist but NONE are active for this combination. "
            f"Found statuses={statuses}. Found terms={terms}. "
            f"Requested term={'none' if not term else term.name!r}. "
            "Set status='active' in the admin panel."
        )

    return _error(
        t("teacher_not_assigned", lang, subject=subject.name, class_level=class_level.name),
        status.HTTP_403_FORBIDDEN, lang=lang, detail=detail,
    )


# =============================================================================
#  Student fetcher
# =============================================================================

def _fetch_students_for_teacher(teacher, academic_year, term, school_level,
                                 class_level, subject):
    ta_qs = TeacherAssignment.objects.filter(
        teacher=teacher, academic_year=academic_year,
        class_level=class_level, subject=subject, status="active",
    )
    if term:
        ta_qs = ta_qs.filter(term=term)

    classroom_ids = (
        ta_qs.values_list("classrooms__id", flat=True)
             .exclude(classrooms__isnull=True)
             .distinct()
    )

    base_qs = Student.objects.filter(
        current_academic_year=academic_year,
        current_school_level=school_level,
        current_class_level=class_level,
        status="active",
    ).order_by("full_name")

    if classroom_ids:
        student_ids = (
            StudentClassroomAssignment.objects.filter(
                academic_year=academic_year,
                classroom_id__in=classroom_ids,
                status="active",
            ).values_list("student_id", flat=True).distinct()
        )
        return base_qs.filter(id__in=student_ids)

    return base_qs


# =============================================================================
#  Grade template builder
# =============================================================================

def _build_grade_template(teacher, academic_year, term, school_level,
                           class_level, subject, grade_type, lang="en") -> bytes:
    students = list(
        _fetch_students_for_teacher(
            teacher, academic_year, term, school_level, class_level, subject
        ).values("id", "roll_number", "full_name")
    )

    grade_type_label = dict(GradeType.choices).get(
        grade_type, grade_type.replace("_", " ").title()
    )
    term_label = term.name if term else "—"

    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Row 1 — banner
    ws.merge_cells("A1:C1")
    _cell(ws, 1, 1, value="STUDENT GRADE RECORDING TEMPLATE",
          font=_WHITE_BOLD, fill=_HDR_FILL, align=_CENTER)
    ws.row_dimensions[1].height = 28

    # Row 2 — school subtitle
    ws.merge_cells("A2:C2")
    _cell(ws, 2, 1, value="École Les Hirondelles de Don Bosco",
          font=Font(name="Arial", italic=True, color="1F4E79", size=10),
          fill=_META_FILL, align=_CENTER)

    # Rows 4-11 — metadata (merge-before-write via helper)
    meta_rows = [
        ("Academic Year",  academic_year.name),
        ("Term",           term_label),
        ("School Level",   school_level.name),
        ("Class Level",    f"{class_level.name} ({class_level.code})"),
        ("Subject",        f"{subject.name} ({subject.code})"),
        ("Grade Type",     grade_type_label),
        ("Teacher",        teacher.full_name),
        ("Total Students", str(len(students))),
    ]
    for i, (label, value) in enumerate(meta_rows, start=4):
        _merged_label_value(
            ws, row=i, label=label, value=value,
            label_font=_DARK_BOLD, label_fill=_META_FILL,
            value_font=_DARK_REG,  value_fill=_META_FILL,
            merge_start_col=2, merge_end_col=3,
        )

    # Row 13 — instructions
    ws.merge_cells("A13:C13")
    _cell(ws, 13, 1,
          value=(
              "⚠ INSTRUCTIONS: Enter scores in the highlighted green column only. "
              "Roll Number and Student Name are locked identifiers — do NOT modify them. "
              f"Valid score range: 0 – 100.  Grade type: {grade_type_label}."
          ),
          font=_RED_BOLD, fill=_WARN_FILL,
          align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[13].height = 36

    # Row 15 — column headers
    for col, hdr in enumerate(
        ["Roll Number", "Student Full Name", f"{grade_type_label}\n(Score 0–100)"],
        start=1,
    ):
        _cell(ws, 15, col, value=hdr, font=_WHITE_BOLD, fill=_COL_FILL, align=_CENTER)
    ws.row_dimensions[15].height = 30

    # Rows 16+ — student data
    for idx, stu in enumerate(students):
        r = 16 + idx
        _cell(ws, r, 1, value=stu["roll_number"], font=_DARK_BOLD, fill=_LOCK_FILL,  align=_CENTER)
        _cell(ws, r, 2, value=stu["full_name"],   font=_DARK_REG,  fill=_LOCK_FILL,  align=_LEFT)
        _cell(ws, r, 3, value=None,               font=_DARK_REG,  fill=_INPUT_FILL, align=_CENTER)

    # Footer
    footer_r = 16 + len(students) + 1
    ws.merge_cells(f"A{footer_r}:C{footer_r}")
    ft = ws.cell(row=footer_r, column=1,
                 value="Grade letter is auto-assigned by the system upon upload.")
    ft.font = _GREY_ITAL; ft.alignment = _LEFT

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A16"

    # Hidden _meta sheet
    m = wb.create_sheet("_meta")
    m["A1"]  = "template_type";      m["B1"]  = "grades"
    m["A2"]  = "academic_year_id";   m["B2"]  = academic_year.id
    m["A3"]  = "academic_year_name"; m["B3"]  = academic_year.name
    m["A4"]  = "term_id";            m["B4"]  = term.id if term else ""
    m["A5"]  = "term_name";          m["B5"]  = term.name if term else ""
    m["A6"]  = "school_level_id";    m["B6"]  = school_level.id
    m["A7"]  = "school_level_name";  m["B7"]  = school_level.name
    m["A8"]  = "class_level_id";     m["B8"]  = class_level.id
    m["A9"]  = "class_level_code";   m["B9"]  = class_level.code
    m["A10"] = "subject_id";         m["B10"] = subject.id
    m["A11"] = "subject_code";       m["B11"] = subject.code
    m["A12"] = "grade_type";         m["B12"] = grade_type
    m["A13"] = "teacher_id";         m["B13"] = teacher.id
    m.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
#  Attendance template builder
# =============================================================================

_ATT_HDR_FILL   = PatternFill("solid", start_color="1E5631")
_ATT_META_FILL  = PatternFill("solid", start_color="D9EAD3")
_ATT_COL_FILL   = PatternFill("solid", start_color="2D6A4F")
_ATT_INPUT_FILL = PatternFill("solid", start_color="D9EAD3")
_ATT_LABEL_FONT = Font(name="Arial", bold=True, color="1E5631", size=10)


def _build_attendance_template(teacher, academic_year, term, school_level,
                                class_level, subject, session_date, lang="en") -> bytes:
    students = list(
        _fetch_students_for_teacher(
            teacher, academic_year, term, school_level, class_level, subject
        ).values("id", "roll_number", "full_name")
    )
    term_label = term.name if term else "—"

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Row 1 — banner
    ws.merge_cells("A1:D1")
    _cell(ws, 1, 1, value="STUDENT ATTENDANCE RECORDING TEMPLATE",
          font=_WHITE_BOLD, fill=_ATT_HDR_FILL, align=_CENTER)
    ws.row_dimensions[1].height = 28

    # Row 2 — school subtitle
    ws.merge_cells("A2:D2")
    _cell(ws, 2, 1, value="École Les Hirondelles de Don Bosco",
          font=Font(name="Arial", italic=True, color="1E5631", size=10),
          fill=_ATT_META_FILL, align=_CENTER)

    # Rows 4-11 — metadata
    meta_rows = [
        ("Academic Year",  academic_year.name),
        ("Term",           term_label),
        ("School Level",   school_level.name),
        ("Class Level",    f"{class_level.name} ({class_level.code})"),
        ("Subject",        f"{subject.name} ({subject.code})"),
        ("Session Date",   str(session_date)),
        ("Teacher",        teacher.full_name),
        ("Total Students", str(len(students))),
    ]
    for i, (label, value) in enumerate(meta_rows, start=4):
        _merged_label_value(
            ws, row=i, label=label, value=value,
            label_font=_ATT_LABEL_FONT, label_fill=_ATT_META_FILL,
            value_font=_DARK_REG,       value_fill=_ATT_META_FILL,
            merge_start_col=2, merge_end_col=4,
        )

    # Row 13 — instructions
    ws.merge_cells("A13:D13")
    _cell(ws, 13, 1,
          value=(
              "⚠ INSTRUCTIONS: Fill the 'Status' column for each student. "
              "Accepted values: present | absent | late | excused (case-insensitive). "
              "Default is 'present'. Do NOT modify Roll Number or Name columns."
          ),
          font=_RED_BOLD, fill=_WARN_FILL,
          align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[13].height = 36

    # Row 15 — column headers
    for col, hdr in enumerate(
        ["Roll Number", "Student Full Name", "Status", "Remarks"], start=1
    ):
        _cell(ws, 15, col, value=hdr, font=_WHITE_BOLD, fill=_ATT_COL_FILL, align=_CENTER)
    ws.row_dimensions[15].height = 22

    # Rows 16+ — student data
    for idx, stu in enumerate(students):
        r = 16 + idx
        _cell(ws, r, 1, value=stu["roll_number"], font=_DARK_BOLD, fill=_LOCK_FILL,       align=_CENTER)
        _cell(ws, r, 2, value=stu["full_name"],   font=_DARK_REG,  fill=_LOCK_FILL,       align=_LEFT)
        _cell(ws, r, 3, value="present",          font=_DARK_REG,  fill=_ATT_INPUT_FILL,  align=_CENTER)
        _cell(ws, r, 4, value="",                 font=_DARK_REG,  fill=_ATT_INPUT_FILL,  align=_LEFT)

    # Footer
    footer_r = 16 + len(students) + 1
    ws.merge_cells(f"A{footer_r}:D{footer_r}")
    ws.cell(row=footer_r, column=1,
            value="Accepted status values: present, absent, late, excused").font = _GREY_ITAL

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.freeze_panes = "A16"

    # Hidden _meta sheet
    m = wb.create_sheet("_meta")
    m["A1"] = "template_type";    m["B1"] = "attendance"
    m["A2"] = "academic_year_id"; m["B2"] = academic_year.id
    m["A3"] = "term_id";          m["B3"] = term.id if term else ""
    m["A4"] = "school_level_id";  m["B4"] = school_level.id
    m["A5"] = "class_level_id";   m["B5"] = class_level.id
    m["A6"] = "subject_id";       m["B6"] = subject.id
    m["A7"] = "session_date";     m["B7"] = str(session_date)
    m["A8"] = "teacher_id";       m["B8"] = teacher.id
    m.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
#  VIEW: download grade template
# =============================================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_grade_template(request):
    """
    Generate and stream a grade recording Excel template.

    Required query params:
      academic_year_id, term_id, school_level_id, class_level_id,
      subject_id, grade_type

    Optional:
      lang  – en | fr | rw  (default: en)
    """
    lang = get_lang(request)

    print(
        f"\n[GRADE TEMPLATE REQUEST]"
        f" user={request.user.username!r} (id={request.user.id})"
        f" | params={dict(request.query_params)}"
        f" | lang={lang}",
        file=sys.stderr, flush=True,
    )

    # Resolve teacher
    if _is_admin(request.user):
        teacher_id = request.query_params.get("teacher_id")
        if not teacher_id:
            return _error(
                "Admin must supply teacher_id to generate a teacher-specific template.",
                status.HTTP_400_BAD_REQUEST, lang=lang,
            )
        try:
            teacher = Teacher.objects.get(pk=teacher_id)
        except Teacher.DoesNotExist as exc:
            return _error(f"Teacher id={teacher_id} not found.",
                          status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    else:
        teacher, err = _get_teacher(request.user, lang)
        if err:
            return err

    # Required params
    ay_id      = request.query_params.get("academic_year_id")
    term_id    = request.query_params.get("term_id")
    sl_id      = request.query_params.get("school_level_id")
    cl_id      = request.query_params.get("class_level_id")
    subj_id    = request.query_params.get("subject_id")
    grade_type = request.query_params.get("grade_type", "").lower()

    missing = [n for n, v in [
        ("academic_year_id", ay_id), ("term_id", term_id),
        ("school_level_id", sl_id),  ("class_level_id", cl_id),
        ("subject_id", subj_id),     ("grade_type", grade_type),
    ] if not v]
    if missing:
        return _error(
            f"Missing required parameters: {', '.join(missing)}",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail="Supply: academic_year_id, term_id, school_level_id, "
                   "class_level_id, subject_id, grade_type",
        )

    valid_grade_types = [c[0] for c in GradeType.choices]
    if grade_type not in valid_grade_types:
        return _error(
            f"Invalid grade_type '{grade_type}'.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=f"Valid options: {', '.join(valid_grade_types)}",
        )

    # Fetch DB objects — each individually so the error says exactly what's missing
    try:
        academic_year = AcademicYear.objects.get(pk=ay_id)
    except AcademicYear.DoesNotExist as exc:
        return _error(f"AcademicYear id={ay_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        term = Term.objects.get(pk=term_id)
    except Term.DoesNotExist as exc:
        return _error(f"Term id={term_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        school_level = SchoolLevel.objects.get(pk=sl_id)
    except SchoolLevel.DoesNotExist as exc:
        return _error(f"SchoolLevel id={sl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        class_level = ClassLevel.objects.select_related("school_level").get(pk=cl_id)
    except ClassLevel.DoesNotExist as exc:
        return _error(f"ClassLevel id={cl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        subject = Subject.objects.get(pk=subj_id)
    except Subject.DoesNotExist as exc:
        return _error(f"Subject id={subj_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)

    # Assignment check (teachers only)
    if not _is_admin(request.user):
        err = _check_teacher_assignment(
            teacher, academic_year, class_level, subject, lang, term=term
        )
        if err:
            return err

    # Build Excel
    try:
        xlsx_bytes = _build_grade_template(
            teacher=teacher, academic_year=academic_year, term=term,
            school_level=school_level, class_level=class_level,
            subject=subject, grade_type=grade_type, lang=lang,
        )
    except Exception as exc:
        return _error("Failed to build the grade template Excel file.",
                      status.HTTP_500_INTERNAL_SERVER_ERROR, lang=lang, exc=exc)

    filename = (
        f"grades_{academic_year.name}_{class_level.code}"
        f"_{subject.code}_{term.name}_{grade_type}.xlsx"
    ).replace("/", "-").replace(" ", "_")

    print(
        f"  [OK] Grade template built: {filename} | {len(xlsx_bytes)} bytes",
        file=sys.stderr, flush=True,
    )

    resp = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["Content-Length"] = len(xlsx_bytes)
    return resp


# =============================================================================
#  VIEW: download attendance template
# =============================================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_attendance_template(request):
    """
    Generate and stream an attendance Excel template.

    Required query params:
      academic_year_id, term_id, school_level_id, class_level_id, subject_id

    Optional:
      session_date  – YYYY-MM-DD (default: today)
      lang          – en | fr | rw
    """
    lang = get_lang(request)

    print(
        f"\n[ATTENDANCE TEMPLATE REQUEST]"
        f" user={request.user.username!r} (id={request.user.id})"
        f" | params={dict(request.query_params)}"
        f" | lang={lang}",
        file=sys.stderr, flush=True,
    )

    # Resolve teacher
    if _is_admin(request.user):
        teacher_id = request.query_params.get("teacher_id")
        if not teacher_id:
            return _error("Admin must supply teacher_id.",
                          status.HTTP_400_BAD_REQUEST, lang=lang)
        try:
            teacher = Teacher.objects.get(pk=teacher_id)
        except Teacher.DoesNotExist as exc:
            return _error(f"Teacher id={teacher_id} not found.",
                          status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    else:
        teacher, err = _get_teacher(request.user, lang)
        if err:
            return err

    # Required params
    ay_id            = request.query_params.get("academic_year_id")
    term_id          = request.query_params.get("term_id")
    sl_id            = request.query_params.get("school_level_id")
    cl_id            = request.query_params.get("class_level_id")
    subj_id          = request.query_params.get("subject_id")
    session_date_str = request.query_params.get("session_date", str(date.today()))

    missing = [n for n, v in [
        ("academic_year_id", ay_id), ("term_id", term_id),
        ("school_level_id", sl_id),  ("class_level_id", cl_id),
        ("subject_id", subj_id),
    ] if not v]
    if missing:
        return _error(f"Missing required parameters: {', '.join(missing)}",
                      status.HTTP_400_BAD_REQUEST, lang=lang)

    try:
        session_date = datetime.strptime(session_date_str, "%Y-%m-%d").date()
    except ValueError as exc:
        return _error("Invalid session_date. Use YYYY-MM-DD format.",
                      status.HTTP_400_BAD_REQUEST, lang=lang, exc=exc,
                      detail=f"Received: '{session_date_str}'")

    # Fetch DB objects
    try:
        academic_year = AcademicYear.objects.get(pk=ay_id)
    except AcademicYear.DoesNotExist as exc:
        return _error(f"AcademicYear id={ay_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        term = Term.objects.get(pk=term_id)
    except Term.DoesNotExist as exc:
        return _error(f"Term id={term_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        school_level = SchoolLevel.objects.get(pk=sl_id)
    except SchoolLevel.DoesNotExist as exc:
        return _error(f"SchoolLevel id={sl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        class_level = ClassLevel.objects.select_related("school_level").get(pk=cl_id)
    except ClassLevel.DoesNotExist as exc:
        return _error(f"ClassLevel id={cl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        subject = Subject.objects.get(pk=subj_id)
    except Subject.DoesNotExist as exc:
        return _error(f"Subject id={subj_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)

    # Assignment check (teachers only)
    if not _is_admin(request.user):
        err = _check_teacher_assignment(
            teacher, academic_year, class_level, subject, lang, term=term
        )
        if err:
            return err

    # Build Excel
    try:
        xlsx_bytes = _build_attendance_template(
            teacher=teacher, academic_year=academic_year, term=term,
            school_level=school_level, class_level=class_level,
            subject=subject, session_date=session_date, lang=lang,
        )
    except Exception as exc:
        return _error("Failed to build the attendance template Excel file.",
                      status.HTTP_500_INTERNAL_SERVER_ERROR, lang=lang, exc=exc)

    filename = (
        f"attendance_{academic_year.name}_{class_level.code}"
        f"_{subject.code}_{session_date_str}.xlsx"
    ).replace("/", "-").replace(" ", "_")

    resp = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["Content-Length"] = len(xlsx_bytes)
    return resp


# =============================================================================
#  VIEW: list trimesters
# =============================================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_trimesters(request):
    """Return the Rwanda trimester list for front-end dropdowns."""
    lang = get_lang(request)
    academic_year_id = request.query_params.get("academic_year_id")
    result = []
    for tri in RWANDA_TRIMESTERS:
        entry = {
            "code":        tri["code"],
            "label":       _tri_label(tri, lang),
            "label_en":    tri["label_en"],
            "label_fr":    tri["label_fr"],
            "label_rw":    tri["label_rw"],
            "month_start": tri["month_start"],
            "month_end":   tri["month_end"],
        }
        if academic_year_id:
            try:
                ay = AcademicYear.objects.get(pk=academic_year_id)
                s, e = _trimester_dates(ay, tri["code"])
                entry["start_date"] = str(s)
                entry["end_date"]   = str(e)
            except AcademicYear.DoesNotExist:
                pass
        result.append(entry)
    return Response({"success": True, "data": result})


# =============================================================================
#  VIEW: upload filled grade template
# =============================================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_grade_template_file(request):
    """
    Process a filled grade template.
    Form fields: excel_file, academic_year_id, term_id, school_level_id,
                 class_level_id, subject_id, grade_type
    """
    from .views import _notify_admins
    from .models import GradeUpload, GradeUploadStatus, StudentGrade
    from decimal import Decimal, InvalidOperation
    from django.db import transaction
    import openpyxl, os

    lang = get_lang(request)

    print(
        f"\n[GRADE UPLOAD REQUEST]"
        f" user={request.user.username!r}"
        f" | fields={list(request.data.keys())}"
        f" | lang={lang}",
        file=sys.stderr, flush=True,
    )

    # ── Auth: teachers only ───────────────────────────────────────────────────
    if _is_admin(request.user):
        return _error(
            "Use the teacher account to upload grades.",
            status.HTTP_403_FORBIDDEN, lang=lang,
        )

    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err

    # ── Required fields ───────────────────────────────────────────────────────
    excel_file = request.FILES.get("excel_file")
    ay_id      = request.data.get("academic_year_id")
    term_id    = request.data.get("term_id")
    sl_id      = request.data.get("school_level_id")
    cl_id      = request.data.get("class_level_id")
    subj_id    = request.data.get("subject_id")
    grade_type = (request.data.get("grade_type") or "").strip().lower()

    missing = [
        name for name, val in [
            ("excel_file",       excel_file),
            ("academic_year_id", ay_id),
            ("term_id",          term_id),
            ("school_level_id",  sl_id),
            ("class_level_id",   cl_id),
            ("subject_id",       subj_id),
            ("grade_type",       grade_type),
        ]
        if not val
    ]
    if missing:
        return _error(
            f"Missing required fields: {', '.join(missing)}",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail="Supply: excel_file, academic_year_id, term_id, "
                   "school_level_id, class_level_id, subject_id, grade_type",
        )

    # ── File extension ────────────────────────────────────────────────────────
    ext = os.path.splitext(excel_file.name)[1].lower()
    if ext not in (".xlsx", ".xls"):
        return _error(
            "Only .xlsx / .xls files are accepted.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=f"Received extension: '{ext}'",
        )

    # ── Grade type validation ─────────────────────────────────────────────────
    valid_grade_types = [c[0] for c in GradeType.choices]
    if grade_type not in valid_grade_types:
        return _error(
            f"Invalid grade_type '{grade_type}'.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=f"Valid options: {', '.join(valid_grade_types)}",
        )

    # ── Fetch DB objects ──────────────────────────────────────────────────────
    try:
        academic_year = AcademicYear.objects.get(pk=ay_id)
    except AcademicYear.DoesNotExist as exc:
        return _error(f"AcademicYear id={ay_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        term = Term.objects.get(pk=term_id)
    except Term.DoesNotExist as exc:
        return _error(f"Term id={term_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        school_level = SchoolLevel.objects.get(pk=sl_id)
    except SchoolLevel.DoesNotExist as exc:
        return _error(f"SchoolLevel id={sl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        class_level = ClassLevel.objects.get(pk=cl_id)
    except ClassLevel.DoesNotExist as exc:
        return _error(f"ClassLevel id={cl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        subject = Subject.objects.get(pk=subj_id)
    except Subject.DoesNotExist as exc:
        return _error(f"Subject id={subj_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)

    # ── Duplicate upload guard ────────────────────────────────────────────────
    if GradeUpload.objects.filter(
        teacher=teacher,
        academic_year=academic_year,
        term=term,
        class_level=class_level,
        subject=subject,
        grade_type=grade_type,
    ).exists():
        return _error(
            "A grade upload already exists for this combination.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=(
                f"teacher_id={teacher.id}, academic_year_id={academic_year.id}, "
                f"term_id={term.id}, class_level_id={class_level.id}, "
                f"subject_id={subject.id}, grade_type={grade_type}"
            ),
        )

    # ── Load workbook ─────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as exc:
        return _error(
            "Cannot read the uploaded Excel file.",
            status.HTTP_400_BAD_REQUEST, lang=lang, exc=exc,
        )

    # ── Locate header row ─────────────────────────────────────────────────────
    #
    # Template layout (fixed, generated by _build_grade_template):
    #   Row  1      : banner        — merged A1:C1
    #   Row  2      : school name   — merged A2:C2
    #   Row  3      : blank
    #   Rows 4–11   : metadata      — label in col A, value merged B:C
    #   Row 12      : blank
    #   Row 13      : instructions  — merged A13:C13
    #                 *** This cell's prose contains BOTH "roll" and "score" ***
    #                 openpyxl returns its value only in column A (index 0);
    #                 columns B and C are None  →  only 1 non-empty cell.
    #   Row 14      : blank
    #   Row 15      : REAL headers  — 3 separate cells:
    #                 A = "Roll Number"
    #                 B = "Student Full Name"
    #                 C = "Quiz\n(Score 0–100)"   ← multi-line via \n
    #   Row 16+     : student data rows
    #
    # Detection rules that accept row 15 and reject every other row:
    #
    #   Rule 1 — at least 2 non-empty cells in the row.
    #            Row 13 (merged instruction) → only col A has a value → FAIL.
    #
    #   Rule 2 — "roll" and "score" must appear in DIFFERENT column indices.
    #            Row 13 → both keywords are in col 0 → col_roll == col_score → FAIL.
    #            Row 15 → col_roll=0 ("Roll Number"), col_score=2 ("Quiz\n(Score…)") → PASS.

    def _normalise(cell_value) -> str:
        """Lowercase, whitespace-collapsed string — handles multi-line cell text."""
        if cell_value is None:
            return ""
        return " ".join(
            str(cell_value).replace("\n", " ").replace("\r", " ").split()
        ).lower()

    header_row_idx: int | None = None
    col_roll:  int | None = None
    col_score: int | None = None

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1
    ):
        if not row:
            continue

        normalised = [_normalise(c) for c in row]

        # Rule 1: need at least 2 non-empty cells.
        if sum(1 for h in normalised if h) < 2:
            continue

        roll_col  = next((i for i, h in enumerate(normalised) if "roll"  in h), None)
        score_col = next((i for i, h in enumerate(normalised) if "score" in h), None)

        if roll_col is None or score_col is None:
            continue

        # Rule 2: keywords must be in different columns.
        if roll_col == score_col:
            print(
                f"  [HEADER SKIP] row={row_idx}"
                f" — 'roll' and 'score' in same column ({roll_col}),"
                f" looks like a merged instruction cell. Skipping.",
                file=sys.stderr, flush=True,
            )
            continue

        # Passed both rules — this is the real header row.
        header_row_idx = row_idx
        col_roll       = roll_col
        col_score      = score_col

        print(
            f"  [HEADER FOUND] row={row_idx}"
            f" | col_roll={col_roll} ('{normalised[col_roll][:40]}')"
            f" | col_score={col_score} ('{normalised[col_score][:40]}')",
            file=sys.stderr, flush=True,
        )
        break

    # ── Fail fast if header never found ──────────────────────────────────────
    if header_row_idx is None:
        dump = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1
        ):
            dump.append(
                f"  row {row_idx:02d}: "
                f"{[str(c)[:50] if c is not None else None for c in (row or [])]}"
            )
        print(
            "  [HEADER DEBUG] Could not locate header row. Sheet dump:\n"
            + "\n".join(dump),
            file=sys.stderr, flush=True,
        )
        return _error(
            "Could not find a valid header row with separate 'Roll Number' and 'Score' columns.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=(
                "Make sure you are uploading the exact template downloaded from this system. "
                "Do not rename, merge, or reorder the column header cells."
            ),
        )

    # ── Build student roster ──────────────────────────────────────────────────
    students_map: dict[str, object] = {
        s.roll_number: s
        for s in _fetch_students_for_teacher(
            teacher, academic_year, term, school_level, class_level, subject
        )
    }

    print(
        f"  [ROSTER] {len(students_map)} students found for"
        f" class={class_level.name!r} subject={subject.name!r}",
        file=sys.stderr, flush=True,
    )

    # ── Parse data rows ───────────────────────────────────────────────────────
    grades_data: list[dict] = []
    row_errors:  list[str]  = []

    for row_num, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        # Skip completely empty rows.
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        # ── Roll number ───────────────────────────────────────────────────────
        roll_raw = row[col_roll] if col_roll < len(row) else None
        roll     = str(roll_raw).strip() if roll_raw is not None else ""

        if not roll:
            continue

        # Skip footer / note rows — real roll numbers are never this long.
        if len(roll) > 30:
            print(
                f"  [SKIP] row={row_num} — too long for a roll number: {roll[:60]!r}",
                file=sys.stderr, flush=True,
            )
            continue

        if roll not in students_map:
            row_errors.append(
                f"Row {row_num}: roll number '{roll}' not found in the class roster."
            )
            continue

        # ── Score ─────────────────────────────────────────────────────────────
        score_raw = row[col_score] if col_score < len(row) else None

        # Blank score → teacher hasn't filled it in yet; skip silently.
        if score_raw is None or str(score_raw).strip() == "":
            continue

        try:
            score = Decimal(str(score_raw).strip())
        except InvalidOperation:
            row_errors.append(
                f"Row {row_num}: cannot convert score value {score_raw!r} to a number."
            )
            continue

        if not (Decimal("0") <= score <= Decimal("100")):
            row_errors.append(
                f"Row {row_num}: score {score} is outside the valid range 0 – 100."
            )
            continue

        grades_data.append({"student": students_map[roll], "score": score})

    print(
        f"  [PARSE] valid_rows={len(grades_data)} row_errors={len(row_errors)}",
        file=sys.stderr, flush=True,
    )
    if row_errors:
        print(
            "  [PARSE ERRORS]\n" + "\n".join(f"    {e}" for e in row_errors),
            file=sys.stderr, flush=True,
        )

    # ── Reject if nothing usable was parsed ──────────────────────────────────
    if not grades_data:
        return _error(
            "No valid grade data found in the uploaded file.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=(
                f"header_row_idx={header_row_idx}, "
                f"col_roll={col_roll}, col_score={col_score}. "
                f"Row errors (first 10): {row_errors[:10]}"
            ),
        )

    # ── Persist ───────────────────────────────────────────────────────────────
    try:
        with transaction.atomic():
            grade_upload = GradeUpload.objects.create(
                teacher=teacher,
                academic_year=academic_year,
                term=term,
                school_level=school_level,
                class_level=class_level,
                subject=subject,
                grade_type=grade_type,
                weight_percentage=GradeType.get_default_weight(grade_type),
                assessment_date=date.today(),
                excel_file=excel_file,
                status=GradeUploadStatus.PENDING,
                uploaded_by=request.user,
            )
            StudentGrade.objects.bulk_create([
                StudentGrade(
                    grade_upload=grade_upload,
                    student=gd["student"],
                    score=gd["score"],
                    max_score=Decimal("100"),
                )
                for gd in grades_data
            ])

        _notify_admins(
            "grade_uploaded",
            t("notif_grade_upload_title", lang),
            t(
                "notif_grade_upload_msg", lang,
                teacher=teacher.full_name,
                subject=subject.name,
                class_level=class_level.name,
            ),
            created_by=request.user,
            extra_data={"grade_upload_id": grade_upload.id},
        )

    except Exception as exc:
        return _error(
            "Database error while saving grades.",
            status.HTTP_500_INTERNAL_SERVER_ERROR, lang=lang, exc=exc,
        )

    print(
        f"  [OK] grade_upload_id={grade_upload.id}"
        f" | records={len(grades_data)}"
        f" | warnings={len(row_errors)}",
        file=sys.stderr, flush=True,
    )

    return Response(
        {
            "success": True,
            "message": (
                f"{len(grades_data)} grade(s) submitted successfully "
                "and are pending admin approval."
            ),
            "data": {
                "grade_upload_id": grade_upload.id,
                "status":          grade_upload.status,
                "grades_saved":    len(grades_data),
                "warnings":        row_errors[:5],
            },
        },
        status=status.HTTP_201_CREATED,
    )
# =============================================================================
#  VIEW: upload filled attendance template
# =============================================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_attendance_template_file(request):
    """
    Process a filled attendance template.
    Form fields: excel_file, academic_year_id, term_id, school_level_id,
                 class_level_id, subject_id, session_date (YYYY-MM-DD, optional)
    """
    from .models import AttendanceSession, StudentAttendance
    from django.db import transaction
    from django.utils import timezone as tz
    import openpyxl, os

    lang = get_lang(request)

    print(
        f"\n[ATTENDANCE UPLOAD REQUEST]"
        f" user={request.user.username!r}"
        f" | fields={list(request.data.keys())}"
        f" | lang={lang}",
        file=sys.stderr, flush=True,
    )

    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err

    excel_file       = request.FILES.get("excel_file")
    ay_id            = request.data.get("academic_year_id")
    term_id          = request.data.get("term_id")
    sl_id            = request.data.get("school_level_id")
    cl_id            = request.data.get("class_level_id")
    subj_id          = request.data.get("subject_id")
    session_date_str = request.data.get("session_date", str(date.today()))

    missing = [n for n, v in [
        ("excel_file", excel_file), ("academic_year_id", ay_id),
        ("term_id", term_id),       ("school_level_id", sl_id),
        ("class_level_id", cl_id),  ("subject_id", subj_id),
    ] if not v]
    if missing:
        return _error(f"Missing: {', '.join(missing)}",
                      status.HTTP_400_BAD_REQUEST, lang=lang)

    try:
        session_date = datetime.strptime(session_date_str, "%Y-%m-%d").date()
    except ValueError as exc:
        return _error("Invalid session_date. Use YYYY-MM-DD.",
                      status.HTTP_400_BAD_REQUEST, lang=lang, exc=exc,
                      detail=f"Received: '{session_date_str}'")

    try:
        academic_year = AcademicYear.objects.get(pk=ay_id)
    except AcademicYear.DoesNotExist as exc:
        return _error(f"AcademicYear id={ay_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        term = Term.objects.get(pk=term_id)
    except Term.DoesNotExist as exc:
        return _error(f"Term id={term_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        school_level = SchoolLevel.objects.get(pk=sl_id)
    except SchoolLevel.DoesNotExist as exc:
        return _error(f"SchoolLevel id={sl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        class_level = ClassLevel.objects.get(pk=cl_id)
    except ClassLevel.DoesNotExist as exc:
        return _error(f"ClassLevel id={cl_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)
    try:
        subject = Subject.objects.get(pk=subj_id)
    except Subject.DoesNotExist as exc:
        return _error(f"Subject id={subj_id} not found.",
                      status.HTTP_404_NOT_FOUND, lang=lang, exc=exc)

    if AttendanceSession.objects.filter(
        teacher=teacher, class_level=class_level,
        subject=subject, session_date=session_date, academic_year=academic_year,
    ).exists():
        return _error(
            "Attendance already recorded for this session.",
            status.HTTP_400_BAD_REQUEST, lang=lang,
            detail=(
                f"teacher_id={teacher.id}, class_level_id={class_level.id}, "
                f"subject_id={subject.id}, session_date={session_date}"
            ),
        )

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as exc:
        return _error("Cannot read the uploaded Excel file.",
                      status.HTTP_400_BAD_REQUEST, lang=lang, exc=exc)

    header_row_idx, headers = None, []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1
    ):
        if row and any(cell and "roll" in str(cell).lower() for cell in row):
            header_row_idx = row_idx
            headers = [str(c).strip().lower() if c else "" for c in row]
            break

    if not headers:
        return _error("Could not find a header row containing 'roll_number'.",
                      status.HTTP_400_BAD_REQUEST, lang=lang)

    col_roll    = next((i for i, h in enumerate(headers) if "roll"   in h), None)
    col_status  = next((i for i, h in enumerate(headers) if "status" in h), None)
    col_remarks = next((i for i, h in enumerate(headers) if "remark" in h), None)

    if col_roll is None:
        return _error("Template must contain a 'roll_number' column.",
                      status.HTTP_400_BAD_REQUEST, lang=lang,
                      detail=f"Detected headers: {headers}")

    students_map = {
        s.roll_number: s
        for s in _fetch_students_for_teacher(
            teacher, academic_year, term, school_level, class_level, subject
        )
    }

    VALID_STATUSES = {"present", "absent", "late", "excused"}
    attendance_data, row_errors = [], []

    for row_num, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or all(c is None or str(c).strip() == "" for c in row[:3]):
            continue
        roll = str(row[col_roll]).strip() if col_roll < len(row) and row[col_roll] else None
        if not roll:
            continue
        if roll not in students_map:
            row_errors.append(f"Row {row_num}: roll '{roll}' not in class roster.")
            continue
        att_status = "present"
        if col_status is not None and col_status < len(row) and row[col_status]:
            raw = str(row[col_status]).strip().lower()
            att_status = raw if raw in VALID_STATUSES else "present"
        remarks = ""
        if col_remarks is not None and col_remarks < len(row) and row[col_remarks]:
            remarks = str(row[col_remarks]).strip()
        attendance_data.append({"student": students_map[roll],
                                 "status": att_status, "remarks": remarks})

    if not attendance_data:
        return _error("No valid attendance data found in the file.",
                      status.HTTP_400_BAD_REQUEST, lang=lang,
                      detail=f"Row errors (first 10): {row_errors[:10]}")

    try:
        with transaction.atomic():
            session = AttendanceSession.objects.create(
                teacher=teacher, academic_year=academic_year, term=term,
                school_level=school_level, class_level=class_level, subject=subject,
                session_date=session_date, excel_file=excel_file,
                is_submitted=True, submitted_at=tz.now(), created_by=request.user,
            )
            for ad in attendance_data:
                StudentAttendance.objects.create(
                    session=session, student=ad["student"],
                    status=ad["status"], remarks=ad["remarks"],
                )
    except Exception as exc:
        return _error("Database error while saving attendance.",
                      status.HTTP_500_INTERNAL_SERVER_ERROR, lang=lang, exc=exc)

    print(
        f"  [OK] Attendance saved: session_id={session.id}"
        f" | records={len(attendance_data)} | warnings={len(row_errors)}",
        file=sys.stderr, flush=True,
    )

    return Response(
        {
            "success": True,
            "message": f"{len(attendance_data)} attendance records saved.",
            "data": {
                "session_id": session.id,
                "date": str(session.session_date),
                "records_saved": len(attendance_data),
                "warnings": row_errors[:5],
            },
        },
        status=status.HTTP_201_CREATED,
    )
    