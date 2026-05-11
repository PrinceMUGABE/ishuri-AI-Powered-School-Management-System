"""
academics_records/utils.py

Utility helpers for the academics_records app.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_grade_template(students, subject_name="", class_level_name="", term=""):
    """
    Generate an Excel template (.xlsx) pre-populated with student roll numbers
    and names, ready for the teacher to fill in scores.

    Returns a BytesIO buffer.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades Template"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["roll_number", "student_name", "score", "max_score", "grade_letter", "remarks"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    # ── Metadata row (row 2 as info) ──────────────────────────────────────────
    info_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    info = ws.cell(row=2, column=1, value=f"Subject: {subject_name} | Class: {class_level_name} | Term: {term}")
    info.fill = info_fill
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # ── Student rows ──────────────────────────────────────────────────────────
    student_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    for row_num, student in enumerate(students, 3):
        ws.cell(row=row_num, column=1, value=student.get("roll_number", "")).fill = student_fill
        ws.cell(row=row_num, column=2, value=student.get("full_name", "")).fill   = student_fill
        ws.cell(row=row_num, column=3, value="").fill   = student_fill   # score – to be filled
        ws.cell(row=row_num, column=4, value=100).fill  = student_fill   # max_score default
        ws.cell(row=row_num, column=5, value="").fill   = student_fill   # grade_letter (optional)
        ws.cell(row=row_num, column=6, value="").fill   = student_fill   # remarks (optional)
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin

    # ── Column widths ─────────────────────────────────────────────────────────
    column_widths = [18, 30, 10, 12, 14, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def compute_discipline_zone(attendance_rate: float) -> str:
    """
    Returns discipline zone string based on attendance rate (0-100).
    low  : ≤ 60%
    medium: 61-80%
    high : > 80%
    """
    if attendance_rate <= 60:
        return 'low'
    elif attendance_rate <= 80:
        return 'medium'
    return 'high'