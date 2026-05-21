# academics_records/views.py

import os
import traceback
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from django.db import models

from django.http import HttpResponse
import openpyxl
from django.db import transaction, IntegrityError
from django.utils import timezone
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from academics.models import AcademicYear, ClassLevel, Subject, Term, SchoolLevel, ClassRoom
from accounts.models import User
from students.models import Student, Parent
from teachers.models import Teacher, TeacherAssignment

from .models import (
    GradeUpload, GradeUploadStatus, GradeType,
    StudentGrade, AttendanceSession, StudentAttendance,
    Assignment
)
from .serializers import (
    AttendanceRecordUpdateSerializer, AttendanceSessionDetailSerializer, GradeUploadSerializer, StudentGradeSerializer, StudentGradeUpdateSerializer,
    StudentAttendanceSerializer, AttendanceSessionSerializer,
    AssignmentSerializer, AssignmentUpdateSerializer

)
from .calculations import GradeCalculator, DisciplineCalculator, PerformanceReportGenerator
from .translations import t, get_lang


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def _ok(data=None, message="", status_code=status.HTTP_200_OK):
    return Response({"success": True, "message": message, "data": data}, status=status_code)


def _err(message="", status_code=status.HTTP_400_BAD_REQUEST, errors=None):
    payload = {"success": False, "message": message}
    if errors:
        payload["errors"] = errors
    return Response(payload, status=status_code)


def _get_teacher(user, lang):
    try:
        teacher = Teacher.objects.get(user=user)
        return teacher, None
    except Teacher.DoesNotExist:
        return None, _err(t("teacher_profile_not_found", lang), status.HTTP_404_NOT_FOUND)


def _is_admin(user):
    return user.role == "admin" or user.is_superuser or user.is_staff


def _auto_grade_letter(score):
    if score >= 90: return 'A+'
    if score >= 80: return 'A'
    if score >= 70: return 'B'
    if score >= 60: return 'C'
    if score >= 50: return 'D'
    return 'F'


def _notify_admins(notification_type, title, message, created_by=None, extra_data=None):
    """Send notification to all admin users"""
    try:
        from notifications.services import NotificationService
        admins = User.objects.filter(role="admin", status="active")
        for admin in admins:
            NotificationService.create_notification(
                recipient=admin,
                notification_type=notification_type,
                title=title,
                message=message,
                priority='medium',
                created_by=created_by,
                data=extra_data or {}
            )
    except Exception as exc:
        print(f"[Notification error] {exc}")


def _notify_user(user, notification_type, title, message, created_by=None, extra_data=None):
    try:
        from notifications.services import NotificationService
        NotificationService.create_notification(
            recipient=user,
            notification_type=notification_type,
            title=title,
            message=message,
            priority='medium',
            created_by=created_by,
            data=extra_data or {}
        )
    except Exception as exc:
        print(f"[Notification error] {exc}")


# ============================================================
# GRADE UPLOAD VIEWS
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_grades(request):
    """Teacher uploads grades Excel file"""
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    # Validate required fields
    academic_year_id = request.data.get('academic_year_id')
    term_id = request.data.get('term_id')
    class_level_id = request.data.get('class_level_id')
    subject_id = request.data.get('subject_id')
    grade_type = request.data.get('grade_type', GradeType.ASSIGNMENT)
    excel_file = request.FILES.get('excel_file')
    
    if not all([academic_year_id, class_level_id, subject_id, excel_file]):
        return _err(t("invalid_data", lang) + " (academic_year_id, class_level_id, subject_id, excel_file required)")
    
    try:
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        term = get_object_or_404(Term, id=term_id) if term_id else None
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        subject = get_object_or_404(Subject, id=subject_id)
        school_level = class_level.school_level
        
        # Verify teacher is assigned to this subject/class
        teacher_assignment = TeacherAssignment.objects.filter(
            teacher=teacher,
            academic_year=academic_year,
            class_level=class_level,
            subject=subject,
            status='active'
        ).first()
        
        if not teacher_assignment:
            return _err(t("teacher_not_assigned", lang, subject=subject.name, class_level=class_level.name), 
                       status.HTTP_403_FORBIDDEN)
        
        # Validate grade type
        if grade_type not in [c[0] for c in GradeType.choices]:
            return _err(f"Invalid grade_type. Choose from: {', '.join([c[0] for c in GradeType.choices])}")
        
        weight = Decimal(str(request.data.get('weight_percentage', GradeType.get_default_weight(grade_type))))
        
        # Check total weight doesn't exceed 100%
        existing_weight = GradeUpload.objects.filter(
            teacher=teacher,
            academic_year=academic_year,
            term=term,
            class_level=class_level,
            subject=subject,
            status=GradeUploadStatus.APPROVED
        ).aggregate(total=models.Sum('weight_percentage'))['total'] or Decimal('0')
        
        if existing_weight + weight > Decimal('100'):
            return _err(f"Total weight would exceed 100% (current: {existing_weight}%, adding: {weight}%)")
        
        # Check for duplicate grade type
        if GradeUpload.objects.filter(
            teacher=teacher, academic_year=academic_year, term=term,
            class_level=class_level, subject=subject, grade_type=grade_type
        ).exists():
            return _err(f"A {grade_type} grade upload already exists for this subject/class/term")
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # Parse Excel file
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Find headers (look for roll_number column)
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            if row and any(cell and 'roll' in str(cell).lower() for cell in row):
                headers = [str(cell).strip().lower() if cell else '' for cell in row]
                break
        
        if not headers:
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        
        col_map = {}
        for idx, h in enumerate(headers):
            if 'roll' in h:
                col_map['roll_number'] = idx
            elif 'score' in h:
                col_map['score'] = idx
            elif 'max' in h:
                col_map['max_score'] = idx
            elif 'remark' in h:
                col_map['remarks'] = idx
        
        if 'roll_number' not in col_map or 'score' not in col_map:
            return _err(t("grade_upload_bad_template", lang, cols="roll_number, score"))
        
        # Get students in this class
        students_in_class = {s.roll_number: s for s in Student.objects.filter(
            current_class_level=class_level,
            current_academic_year=academic_year,
            status='active'
        )}
        
        # Parse grades
        grades_data = []
        invalid_students = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row[:3]):
                continue
            
            roll_number = str(row[col_map['roll_number']]).strip() if col_map['roll_number'] < len(row) and row[col_map['roll_number']] else None
            if not roll_number:
                continue
            
            if roll_number not in students_in_class:
                invalid_students.append(roll_number)
                continue
            
            score_raw = row[col_map['score']] if col_map['score'] < len(row) else None
            if score_raw is None:
                continue
            
            try:
                score = Decimal(str(score_raw))
                if score < 0 or score > 100:
                    return _err(f"Row {row_num}: Score must be between 0 and 100")
            except InvalidOperation:
                return _err(f"Row {row_num}: Invalid score value '{score_raw}'")
            
            max_score = Decimal('100')
            if 'max_score' in col_map and col_map['max_score'] < len(row) and row[col_map['max_score']]:
                try:
                    max_score = Decimal(str(row[col_map['max_score']]))
                except:
                    pass
            
            remarks = ""
            if 'remarks' in col_map and col_map['remarks'] < len(row) and row[col_map['remarks']]:
                remarks = str(row[col_map['remarks']])
            
            grades_data.append({
                'student': students_in_class[roll_number],
                'score': score,
                'max_score': max_score,
                'remarks': remarks
            })
        
        if invalid_students:
            return _err(t("students_not_in_class", lang, students=", ".join(invalid_students[:5]), class_level=class_level.name))
        
        if not grades_data:
            return _err(t("grade_upload_no_data", lang))
        
    except Exception as exc:
        return _err(t("parse_error", lang, error=str(exc)))
    
    # Save grade upload and grades
    try:
        with transaction.atomic():
            grade_upload = GradeUpload.objects.create(
                teacher=teacher,
                academic_year=academic_year,
                term=term,
                school_level=school_level,
                class_level=class_level,
                subject=subject,
                classroom=teacher_assignment.classrooms.first(),
                grade_type=grade_type,
                weight_percentage=weight,
                assessment_date=request.data.get('assessment_date') or date.today(),
                excel_file=excel_file,
                status=GradeUploadStatus.PENDING,
                uploaded_by=request.user
            )
            
            for gd in grades_data:
                StudentGrade.objects.create(
                    grade_upload=grade_upload,
                    student=gd['student'],
                    score=gd['score'],
                    max_score=gd['max_score'],
                    remarks=gd['remarks']
                )
            
            # Notify admins
            _notify_admins(
                'grade_uploaded',
                t("notif_grade_upload_title", lang),
                t("notif_grade_upload_msg", lang, teacher=teacher.full_name, 
                  subject=subject.name, class_level=class_level.name),
                created_by=request.user,
                extra_data={'grade_upload_id': grade_upload.id}
            )
            
        return _ok(data={'id': grade_upload.id, 'status': grade_upload.status}, 
                   message=t("grade_upload_success", lang), 
                   status_code=status.HTTP_201_CREATED)
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_grade_upload(request, upload_id):
    """Admin approves a grade upload"""
    lang = get_lang(request)
    
    if not _is_admin(request.user):
        return _err(t("forbidden", lang), status.HTTP_403_FORBIDDEN)
    
    grade_upload = get_object_or_404(GradeUpload, id=upload_id)
    
    if grade_upload.status == GradeUploadStatus.APPROVED:
        return _err("Grade upload already approved", status.HTTP_400_BAD_REQUEST)
    
    action = request.data.get('action')
    rejection_reason = request.data.get('rejection_reason', '')
    
    with transaction.atomic():
        if action == 'approve':
            grade_upload.status = GradeUploadStatus.APPROVED
            grade_upload.student_grades.update(is_published=True, published_at=timezone.now())
            message = t("grade_approved", lang)
        else:
            grade_upload.status = GradeUploadStatus.REJECTED
            grade_upload.rejection_reason = rejection_reason
            message = t("grade_rejected", lang)
        
        grade_upload.reviewed_by = request.user
        grade_upload.reviewed_at = timezone.now()
        grade_upload.save()
        
        # Notify teacher
        _notify_user(
            grade_upload.teacher.user,
            'grade_reviewed',
            f"Grade Upload {action}d",
            f"Your grade upload for {grade_upload.subject.name} has been {action}d." +
            (f" Reason: {rejection_reason}" if rejection_reason else ""),
            created_by=request.user
        )
    
    return _ok(message=message)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_grade_uploads(request):
    """Get grade uploads (filtered by role)"""
    lang = get_lang(request)
    
    if _is_admin(request.user):
        uploads = GradeUpload.objects.all().select_related('teacher', 'subject', 'class_level', 'academic_year')
    else:
        teacher, err = _get_teacher(request.user, lang)
        if err:
            return err
        uploads = GradeUpload.objects.filter(teacher=teacher).select_related('subject', 'class_level', 'academic_year')
    
    # Apply filters
    status_filter = request.query_params.get('status')
    if status_filter:
        uploads = uploads.filter(status=status_filter)
    
    class_level_id = request.query_params.get('class_level_id')
    if class_level_id:
        uploads = uploads.filter(class_level_id=class_level_id)
    
    subject_id = request.query_params.get('subject_id')
    if subject_id:
        uploads = uploads.filter(subject_id=subject_id)
    
    academic_year_id = request.query_params.get('academic_year_id')
    if academic_year_id:
        uploads = uploads.filter(academic_year_id=academic_year_id)
    
    # Serialize
    result = []
    for u in uploads:
        result.append({
            'id': u.id,
            'teacher': u.teacher.full_name,
            'academic_year': u.academic_year.name,
            'term': u.term.name if u.term else None,
            'class_level': u.class_level.name,
            'subject': u.subject.name,
            'grade_type': u.get_grade_type_display(),
            'weight_percentage': float(u.weight_percentage),
            'status': u.status,
            'created_at': u.created_at.isoformat(),
            'grades_count': u.student_grades.count()
        })
    
    return _ok(data=result, message=t("grade_upload_list_fetched", lang))


# ============================================================
# PERFORMANCE VIEWS (REAL-TIME CALCULATIONS)
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_performance(request, student_id=None):
    """
    Get student performance - ALL CALCULATED IN REAL-TIME.
    No data stored in database - purely computed from source data.
    """
    lang = get_lang(request)
    user = request.user
    
    try:
        # Determine which student to view with permission checks
        if student_id:
            if user.role == 'admin':
                student = get_object_or_404(Student, id=student_id)
            elif user.role == 'teacher':
                student = get_object_or_404(Student, id=student_id)
                teacher = get_object_or_404(Teacher, user=user)
                # Verify teacher teaches this student
                teaches = TeacherAssignment.objects.filter(
                    teacher=teacher,
                    class_level=student.current_class_level,
                    status='active'
                ).exists()
                if not teaches:
                    return _err(t("forbidden", lang), status.HTTP_403_FORBIDDEN)
            elif user.role == 'parent':
                parent = get_object_or_404(Parent, user=user)
                if not parent.students.filter(id=student_id).exists():
                    return _err(t("forbidden", lang), status.HTTP_403_FORBIDDEN)
                student = get_object_or_404(Student, id=student_id)
            else:
                return _err(t("forbidden", lang), status.HTTP_403_FORBIDDEN)
        else:
            # Self view
            if hasattr(user, 'student_profile'):
                student = user.student_profile
            else:
                return _err("No student profile found", status.HTTP_404_NOT_FOUND)
        
        # Get query parameters
        academic_year_id = request.query_params.get('academic_year_id')
        term_id = request.query_params.get('term_id')
        
        if not academic_year_id:
            academic_year = AcademicYear.objects.filter(is_current=True).first()
            if not academic_year:
                return _err("No academic year specified and no current year found")
        else:
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        term = get_object_or_404(Term, id=term_id) if term_id else None
        
        # Generate complete report in real-time
        report = PerformanceReportGenerator.get_full_report(student, academic_year, term)
        
        return _ok(data=report, message=t("performance_fetched", lang))
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_subject_performance(request, student_id, subject_id):
    """Get detailed subject performance - REAL-TIME calculation"""
    lang = get_lang(request)
    
    try:
        student = get_object_or_404(Student, id=student_id)
        subject = get_object_or_404(Subject, id=subject_id)
        
        academic_year_id = request.query_params.get('academic_year_id')
        term_id = request.query_params.get('term_id')
        
        if not academic_year_id:
            academic_year = AcademicYear.objects.filter(is_current=True).first()
        else:
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        term = get_object_or_404(Term, id=term_id) if term_id else None
        
        # Calculate in real-time
        grade_result = GradeCalculator.get_subject_result(student, subject, academic_year, term)
        discipline_result = DisciplineCalculator.get_subject_discipline(student, subject, academic_year, term)
        
        return _ok(data={
            'student': {
                'id': student.id,
                'name': student.full_name,
                'roll_number': student.roll_number
            },
            'subject': {
                'id': subject.id,
                'name': subject.name,
                'code': subject.code,
                'pass_mark': float(subject.pass_mark)
            },
            'academic_year': academic_year.name,
            'term': term.name if term else None,
            'grades': grade_result,
            'discipline': discipline_result,
            'generated_at': timezone.now().isoformat()
        }, message=t("performance_fetched", lang))
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_class_performance(request, class_level_id):
    """Get performance summary for all students in a class"""
    lang = get_lang(request)
    
    try:
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        
        academic_year_id = request.query_params.get('academic_year_id')
        term_id = request.query_params.get('term_id')
        
        if not academic_year_id:
            academic_year = AcademicYear.objects.filter(is_current=True).first()
        else:
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        term = get_object_or_404(Term, id=term_id) if term_id else None
        
        students = Student.objects.filter(
            current_class_level=class_level,
            current_academic_year=academic_year,
            status='active'
        )
        
        results = []
        for student in students:
            performance = GradeCalculator.get_overall_performance(student, academic_year, term)
            discipline = DisciplineCalculator.get_attendance_summary(student, academic_year, term)
            
            results.append({
                'student_id': student.id,
                'student_name': student.full_name,
                'roll_number': student.roll_number,
                'overall_average': performance['overall_average'],
                'grade_letter': performance['grade_letter'],
                'subjects_passed': performance['subjects_passed'],
                'subjects_failed': performance['subjects_failed'],
                'attendance_rate': discipline['attendance_rate'],
                'discipline_score': discipline['discipline_score']
            })
        
        # Sort by overall average
        results.sort(key=lambda x: x['overall_average'] or 0, reverse=True)
        
        # Add rank
        for idx, result in enumerate(results, 1):
            result['rank'] = idx
        
        return _ok(data={
            'class_level': class_level.name,
            'academic_year': academic_year.name,
            'term': term.name if term else None,
            'total_students': len(results),
            'students': results
        }, message=t("performance_fetched", lang))
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================
# ATTENDANCE VIEWS
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_attendance(request):
    """Teacher uploads attendance Excel file"""
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    academic_year_id = request.data.get('academic_year_id')
    class_level_id = request.data.get('class_level_id')
    subject_id = request.data.get('subject_id')
    session_date = request.data.get('session_date', str(date.today()))
    excel_file = request.FILES.get('excel_file')
    
    if not all([academic_year_id, class_level_id, subject_id, excel_file]):
        return _err(t("invalid_data", lang))
    
    try:
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        subject = get_object_or_404(Subject, id=subject_id)
        session_date = datetime.strptime(session_date, '%Y-%m-%d').date()
        
        # Check for duplicate
        if AttendanceSession.objects.filter(
            teacher=teacher, class_level=class_level, 
            subject=subject, session_date=session_date
        ).exists():
            return _err("Attendance already recorded for this date")
        
        # Parse Excel file
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Find headers
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        col_map = {}
        for idx, h in enumerate(headers):
            if 'roll' in h:
                col_map['roll_number'] = idx
            elif 'status' in h:
                col_map['status'] = idx
            elif 'remark' in h:
                col_map['remarks'] = idx
        
        if 'roll_number' not in col_map:
            return _err(t("attendance_upload_bad_template", lang, cols="roll_number, status"))
        
        # Get students
        students_map = {s.roll_number: s for s in Student.objects.filter(
            current_class_level=class_level, current_academic_year=academic_year, status='active'
        )}
        
        attendance_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            roll_number = str(row[col_map['roll_number']]).strip() if row[col_map['roll_number']] else None
            if not roll_number or roll_number not in students_map:
                continue
            
            status_raw = str(row[col_map.get('status', 1)]).strip().lower() if col_map.get('status') and row[col_map.get('status')] else 'present'
            status_map = {'present': 'present', 'absent': 'absent', 'late': 'late', 'excused': 'excused', '1': 'present', '0': 'absent', 'true': 'present', 'false': 'absent', 'yes': 'present', 'no': 'absent'}
            status = status_map.get(status_raw, 'present')
            
            remarks = str(row[col_map.get('remarks', 2)]).strip() if col_map.get('remarks') and row[col_map.get('remarks')] else ''
            
            attendance_data.append({
                'student': students_map[roll_number],
                'status': status,
                'remarks': remarks
            })
        
        if not attendance_data:
            return _err(t("attendance_upload_no_data", lang))
        
        # Save session and records
        with transaction.atomic():
            session = AttendanceSession.objects.create(
                teacher=teacher,
                academic_year=academic_year,
                school_level=class_level.school_level,
                class_level=class_level,
                subject=subject,
                session_date=session_date,
                excel_file=excel_file,
                is_submitted=True,
                submitted_at=timezone.now(),
                created_by=request.user
            )
            
            for ad in attendance_data:
                StudentAttendance.objects.create(
                    session=session,
                    student=ad['student'],
                    status=ad['status'],
                    remarks=ad['remarks']
                )
        
        return _ok(data={'session_id': session.id, 'records': len(attendance_data)}, 
                   message="Attendance uploaded successfully", status_code=status.HTTP_201_CREATED)
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_attendance(request, student_id):
    """Get attendance summary for a student"""
    lang = get_lang(request)
    
    try:
        student = get_object_or_404(Student, id=student_id)
        
        academic_year_id = request.query_params.get('academic_year_id')
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id) if academic_year_id else AcademicYear.objects.filter(is_current=True).first()
        
        summary = DisciplineCalculator.get_attendance_summary(student, academic_year)
        
        # Get detailed records
        records = StudentAttendance.objects.filter(
            student=student,
            session__academic_year=academic_year
        ).select_related('session__subject', 'session__teacher').order_by('-session__session_date')
        
        records_data = []
        for r in records:
            records_data.append({
                'date': r.session.session_date.isoformat(),
                'subject': r.session.subject.name,
                'teacher': r.session.teacher.full_name,
                'status': r.status,
                'remarks': r.remarks
            })
        
        return _ok(data={
            'summary': summary,
            'records': records_data
        })
        
    except Exception as exc:
        return _err(t("unexpected_error", lang, error=str(exc)), status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================
# ASSIGNMENT VIEWS
# ============================================================
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_assignment(request):
    """
    Teacher uploads an assignment PDF
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    # Validate required fields
    title = request.data.get('title')
    academic_year_id = request.data.get('academic_year_id')
    class_level_id = request.data.get('class_level_id')
    subject_id = request.data.get('subject_id')
    pdf_file = request.FILES.get('pdf_file')
    
    if not all([title, academic_year_id, class_level_id, subject_id, pdf_file]):
        return _err("Missing required fields: title, academic_year_id, class_level_id, subject_id, pdf_file", 
                   status.HTTP_400_BAD_REQUEST)
    
    # Validate file type
    if not pdf_file.name.endswith('.pdf'):
        return _err("Only PDF files are allowed", status.HTTP_400_BAD_REQUEST)
    
    try:
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        subject = get_object_or_404(Subject, id=subject_id)
        school_level = class_level.school_level
        
        term_id = request.data.get('term_id')
        term = get_object_or_404(Term, id=term_id) if term_id else None
        
        classroom_id = request.data.get('classroom_id')
        classroom = get_object_or_404(ClassRoom, id=classroom_id) if classroom_id else None
        
        # Verify teacher is assigned to this subject/class
        is_assigned = TeacherAssignment.objects.filter(
            teacher=teacher,
            academic_year=academic_year,
            class_level=class_level,
            subject=subject,
            status='active'
        ).exists()
        
        if not is_assigned:
            return _err(f"You are not assigned to teach {subject.name} in {class_level.name}", 
                       status.HTTP_403_FORBIDDEN)
        
        # Create assignment
        assignment = Assignment.objects.create(
            teacher=teacher,
            academic_year=academic_year,
            term=term,
            school_level=school_level,
            class_level=class_level,
            subject=subject,
            classroom=classroom,
            title=title,
            description=request.data.get('description', ''),
            instructions=request.data.get('instructions', ''),
            pdf_file=pdf_file,
            due_date=request.data.get('due_date') or None,
            due_time=request.data.get('due_time') or None,
            total_marks=request.data.get('total_marks') or None,
            status=Assignment.AssignmentStatus.ACTIVE,
            uploaded_by=request.user
        )
        
        # Notify students in this class - FIXED RELATIONSHIP NAME
        from students.models import StudentClassroomAssignment
        students = Student.objects.filter(
            classroom_assignments__class_level=class_level,
            classroom_assignments__academic_year=academic_year,
            classroom_assignments__status='active',
            status='active'
        ).distinct()
        
        notified_count = 0
        for student in students:
            if student.user:
                try:
                    _notify_user(
                        student.user,
                        'assignment_created',
                        f"New Assignment: {title}",
                        f"A new assignment has been posted for {subject.name}. Due: {assignment.due_date or 'Not specified'}",
                        created_by=request.user,
                        extra_data={'assignment_id': assignment.id, 'subject': subject.name}
                    )
                    notified_count += 1
                except Exception as notify_error:
                    print(f"Failed to notify student {student.id}: {notify_error}")
        
        return _ok(
            data={
                'id': assignment.id,
                'title': assignment.title,
                'pdf_url': assignment.pdf_file.url if assignment.pdf_file else None,
                'notified_students': notified_count
            },
            message=f"Assignment uploaded successfully. {notified_count} students notified.",
            status_code=status.HTTP_201_CREATED
        )
        
    except Exception as exc:
        print(f"Error in upload_assignment: {exc}")
        import traceback
        traceback.print_exc()
        return _err(f"An unexpected error occurred: {str(exc)}", status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_assignments(request):
    """
    Get assignments based on user role
    """
    lang = get_lang(request)
    user = request.user
    
    try:
        if user.role == 'admin':
            # Admin can see all assignments
            assignments = Assignment.objects.all().select_related(
                'teacher', 'subject', 'class_level', 'academic_year'
            ).order_by('-created_at')
            
        elif user.role == 'teacher':
            # Teacher sees their own assignments
            teacher = get_object_or_404(Teacher, user=user)
            assignments = Assignment.objects.filter(teacher=teacher).select_related(
                'subject', 'class_level', 'academic_year'
            ).order_by('-created_at')
            
        elif user.role == 'student':
            # Student sees assignments for their class
            student = user.student_profile
            assignments = Assignment.objects.filter(
                class_level=student.current_class_level,
                status='active'
            ).select_related('teacher', 'subject').order_by('-created_at')
            
        else:
            return _err("Invalid user role", status.HTTP_403_FORBIDDEN)
        
        # Apply filters
        academic_year_id = request.query_params.get('academic_year_id')
        if academic_year_id:
            assignments = assignments.filter(academic_year_id=academic_year_id)
        
        term_id = request.query_params.get('term_id')
        if term_id:
            assignments = assignments.filter(term_id=term_id)
        
        class_level_id = request.query_params.get('class_level_id')
        if class_level_id:
            assignments = assignments.filter(class_level_id=class_level_id)
        
        subject_id = request.query_params.get('subject_id')
        if subject_id:
            assignments = assignments.filter(subject_id=subject_id)
        
        classroom_id = request.query_params.get('classroom_id')
        if classroom_id:
            assignments = assignments.filter(classroom_id=classroom_id)
        
        # Serialize
        from .serializers import AssignmentSerializer
        serializer = AssignmentSerializer(assignments, many=True)
        
        return _ok(
            data=serializer.data,
            message="Assignments fetched successfully"
        )
        
    except Exception as exc:
        print(f"Error in get_assignments: {exc}")
        import traceback
        traceback.print_exc()
        return _err(f"An unexpected error occurred: {str(exc)}", status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def assignment_detail(request, assignment_id):
    """
    Get, update, or delete a specific assignment
    """
    lang = get_lang(request)
    user = request.user
    
    # Determine access
    if user.role == 'admin':
        assignment = get_object_or_404(Assignment, id=assignment_id)
    elif user.role == 'teacher':
        teacher = get_object_or_404(Teacher, user=user)
        assignment = get_object_or_404(Assignment, id=assignment_id, teacher=teacher)
    else:
        return _err("Permission denied", status.HTTP_403_FORBIDDEN)
    
    if request.method == 'GET':
        from .serializers import AssignmentSerializer
        serializer = AssignmentSerializer(assignment)
        return _ok(data=serializer.data, message="Assignment details fetched")
    
    elif request.method == 'PATCH':
        # Update assignment
        from .serializers import AssignmentUpdateSerializer
        serializer = AssignmentUpdateSerializer(assignment, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            
            # Notify students if due date changed
            if 'due_date' in request.data and assignment.class_level:
                students = Student.objects.filter(
                    current_class_level=assignment.class_level,
                    status='active'
                )
                for student in students:
                    if student.user:
                        try:
                            _notify_user(
                                student.user,
                                'assignment_updated',
                                f"Assignment Updated: {assignment.title}",
                                f"The due date for {assignment.subject.name} assignment has been updated to {assignment.due_date}",
                                created_by=request.user,
                                extra_data={'assignment_id': assignment.id}
                            )
                        except Exception:
                            pass
            
            return _ok(data=serializer.data, message="Assignment updated successfully")
        return _err(serializer.errors, status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        assignment_title = assignment.title
        assignment.delete()
        return _ok(message=f"Assignment '{assignment_title}' deleted successfully")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_assignment_file(request, assignment_id):
    """
    Download the PDF file for an assignment
    """
    lang = get_lang(request)
    user = request.user
    
    # Check permissions
    if user.role == 'admin':
        assignment = get_object_or_404(Assignment, id=assignment_id)
    elif user.role == 'teacher':
        teacher = get_object_or_404(Teacher, user=user)
        assignment = get_object_or_404(Assignment, id=assignment_id, teacher=teacher)
    elif user.role == 'student':
        student = user.student_profile
        assignment = get_object_or_404(Assignment, id=assignment_id, class_level=student.current_class_level)
    else:
        return _err("Permission denied", status.HTTP_403_FORBIDDEN)
    
    if not assignment.pdf_file or not assignment.pdf_file.path:
        return _err("File not found", status.HTTP_404_NOT_FOUND)
    
    try:
        from django.http import FileResponse
        import mimetypes
        
        # Open the file and return as response
        response = FileResponse(
            open(assignment.pdf_file.path, 'rb'),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'inline; filename="{assignment.pdf_file.name}"'
        return response
        
    except Exception as exc:
        return _err(f"Error reading file: {str(exc)}", status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def preview_assignment_file(request, assignment_id):
    """
    Preview the PDF file for an assignment (returns URL or base64)
    """
    lang = get_lang(request)
    user = request.user
    
    # Check permissions
    if user.role == 'admin':
        assignment = get_object_or_404(Assignment, id=assignment_id)
    elif user.role == 'teacher':
        teacher = get_object_or_404(Teacher, user=user)
        assignment = get_object_or_404(Assignment, id=assignment_id, teacher=teacher)
    elif user.role == 'student':
        student = user.student_profile
        assignment = get_object_or_404(Assignment, id=assignment_id, class_level=student.current_class_level)
    else:
        return _err("Permission denied", status.HTTP_403_FORBIDDEN)
    
    if not assignment.pdf_file or not assignment.pdf_file.url:
        return _err("File not found", status.HTTP_404_NOT_FOUND)
    
    # For simplicity, return the URL to the PDF file
    return _ok(data={'pdf_url': assignment.pdf_file.url}, message="Assignment preview URL fetched successfully")    

# ============================================================
# TEACHER STUDENTS VIEW
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_students(request):
    """Get students for teacher's current classes"""
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    academic_year_id = request.query_params.get('academic_year_id')
    academic_year = get_object_or_404(AcademicYear, id=academic_year_id) if academic_year_id else AcademicYear.objects.filter(is_current=True).first()
    
    # Get teacher's assignments
    assignments = TeacherAssignment.objects.filter(
        teacher=teacher,
        academic_year=academic_year,
        status='active'
    ).select_related('class_level')
    
    result = []
    for assignment in assignments:
        students = Student.objects.filter(
            current_class_level=assignment.class_level,
            current_academic_year=academic_year,
            status='active'
        ).values('id', 'full_name', 'roll_number')
        
        result.append({
            'class_level_id': assignment.class_level.id,
            'class_level_name': assignment.class_level.name,
            'subject_id': assignment.subject.id,
            'subject_name': assignment.subject.name,
            'students': list(students)
        })
    
    return _ok(data=result, message="Students fetched successfully")



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_grades(request):
    """
    Get all student grades with filtering options.
    Supports filtering by grade_upload_id, student_id, etc.
    """
    lang = get_lang(request)
    
    # Permission check - only admins and teachers can access
    user = request.user
    if user.role not in ['admin', 'teacher']:
        return _err(t("permission_denied", lang), status.HTTP_403_FORBIDDEN, lang=lang)
    
    # Base queryset
    grades = StudentGrade.objects.select_related(
        'grade_upload', 'student', 'grade_upload__academic_year',
        'grade_upload__term', 'grade_upload__class_level',
        'grade_upload__subject', 'grade_upload__school_level',
        'grade_upload__teacher'
    )
    
    # Filter by grade_upload_id (for viewing grades of a specific upload)
    grade_upload_id = request.query_params.get('grade_upload_id')
    if grade_upload_id:
        grades = grades.filter(grade_upload_id=grade_upload_id)
    
    # Filter by student
    student_id = request.query_params.get('student_id')
    if student_id:
        grades = grades.filter(student_id=student_id)
    
    # Filter by academic year
    academic_year_id = request.query_params.get('academic_year_id')
    if academic_year_id:
        grades = grades.filter(grade_upload__academic_year_id=academic_year_id)
    
    # Filter by term
    term_id = request.query_params.get('term_id')
    if term_id:
        grades = grades.filter(grade_upload__term_id=term_id)
    
    # Filter by class level
    class_level_id = request.query_params.get('class_level_id')
    if class_level_id:
        grades = grades.filter(grade_upload__class_level_id=class_level_id)
    
    # Filter by subject
    subject_id = request.query_params.get('subject_id')
    if subject_id:
        grades = grades.filter(grade_upload__subject_id=subject_id)
    
    # Filter by grade type
    grade_type = request.query_params.get('grade_type')
    if grade_type:
        grades = grades.filter(grade_upload__grade_type=grade_type)
    
    # For teachers, only show grades from their uploads
    if user.role == 'teacher':
        try:
            teacher = Teacher.objects.get(user=user)
            grades = grades.filter(grade_upload__teacher=teacher)
        except Teacher.DoesNotExist:
            return _err("Teacher profile not found", status.HTTP_404_NOT_FOUND, lang=lang)
    
    # Order by most recent first
    grades = grades.order_by('-created_at')
    
    # Pagination
    page = int(request.query_params.get('page', 1))
    page_size = min(int(request.query_params.get('page_size', 10)), 100)
    total = grades.count()
    start = (page - 1) * page_size
    paginated_grades = grades[start:start + page_size]
    
    # Serialize
    result = []
    for grade in paginated_grades:
        upload = grade.grade_upload
        percentage = (grade.score / grade.max_score * 100) if grade.max_score > 0 else 0
        
        result.append({
            'id': grade.id,
            'student_id': grade.student.id,
            'student_name': grade.student.full_name,
            'student_roll': grade.student.roll_number,
            'academic_year_id': upload.academic_year.id,
            'academic_year_name': upload.academic_year.name,
            'term_id': upload.term.id if upload.term else None,
            'term_name': upload.term.name if upload.term else None,
            'school_level_id': upload.school_level.id,
            'school_level_name': upload.school_level.name,
            'class_level_id': upload.class_level.id,
            'class_level_name': upload.class_level.name,
            'subject_id': upload.subject.id,
            'subject_name': upload.subject.name,
            'grade_type': upload.grade_type,
            'grade_type_display': upload.get_grade_type_display(),
            'score': float(grade.score),
            'max_score': float(grade.max_score),
            'percentage': round(percentage, 2),
            'custom_grade_letter': grade.custom_grade_letter,
            'remarks': grade.remarks,
            'is_published': grade.is_published,
            'published_at': grade.published_at.isoformat() if grade.published_at else None,
            'grade_upload_id': upload.id,
            'upload_status': upload.status,
            'teacher_name': upload.teacher.full_name if upload.teacher else 'Admin',
            'created_at': grade.created_at.isoformat(),
            'updated_at': grade.updated_at.isoformat(),
        })
    
    return Response({
        'success': True,
        'message': t('grades_fetched', lang),
        'data': {
            'results': result,
            'count': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })

@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def student_grade_detail(request, grade_id):
    """
    Get, update, or delete a specific student grade.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    try:
        grade = StudentGrade.objects.get(
            id=grade_id,
            grade_upload__teacher=teacher
        )
    except StudentGrade.DoesNotExist:
        return _err("Grade not found", status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        upload = grade.grade_upload
        percentage = (grade.score / grade.max_score * 100) if grade.max_score > 0 else 0
        
        data = {
            'id': grade.id,
            'student_id': grade.student.id,
            'student_name': grade.student.full_name,
            'student_roll': grade.student.roll_number,
            'score': float(grade.score),
            'max_score': float(grade.max_score),
            'percentage': round(percentage, 2),
            'custom_grade_letter': grade.custom_grade_letter,
            'remarks': grade.remarks,
            'is_published': grade.is_published,
            'grade_upload_id': upload.id,
        }
        return _ok(data=data)
    
    elif request.method == 'PATCH':
        # Update grade
        score = request.data.get('score')
        max_score = request.data.get('max_score')
        remarks = request.data.get('remarks', '')
        custom_grade_letter = request.data.get('custom_grade_letter', '')
        
        if score is not None:
            try:
                grade.score = Decimal(str(score))
            except (InvalidOperation, TypeError):
                return _err("Invalid score value")
        
        if max_score is not None:
            try:
                grade.max_score = Decimal(str(max_score))
            except (InvalidOperation, TypeError):
                return _err("Invalid max_score value")
        
        grade.remarks = remarks
        grade.custom_grade_letter = custom_grade_letter
        grade.save()
        
        # If grade upload is still pending, it remains pending
        # If approved, we might want to flag that changes were made
        if grade.grade_upload.status == GradeUploadStatus.APPROVED:
            # Optionally notify admin that a grade was modified after approval
            _notify_admins(
                'grade_modified',
                "Grade Modified After Approval",
                f"Teacher {teacher.full_name} modified grade for {grade.student.full_name} in {grade.grade_upload.subject.name}",
                created_by=request.user,
                extra_data={'grade_id': grade.id}
            )
        
        return _ok(message="Grade updated successfully")
    
    elif request.method == 'DELETE':
        upload = grade.grade_upload
        grade.delete()
        
        # Log deletion
        _notify_admins(
            'grade_deleted',
            "Grade Deleted",
            f"Teacher {teacher.full_name} deleted grade for student in {upload.subject.name}",
            created_by=request.user,
            extra_data={'subject': upload.subject.name, 'teacher': teacher.full_name}
        )
        
        return _ok(message="Grade deleted successfully")


@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def grade_upload_detail(request, upload_id):
    """
    Get details of a specific grade upload or delete it (with all associated grades).
    """
    lang = get_lang(request)
    
    # Permission check
    user = request.user
    if user.role not in ['admin', 'teacher']:
        return _err(t("permission_denied", lang), status.HTTP_403_FORBIDDEN, lang=lang)
    
    try:
        if user.role == 'admin':
            upload = get_object_or_404(GradeUpload, id=upload_id)
        else:
            teacher, err = _get_teacher(request.user, lang)
            if err:
                return err
            upload = get_object_or_404(GradeUpload, id=upload_id, teacher=teacher)
    except GradeUpload.DoesNotExist:
        return _err("Grade upload not found", status.HTTP_404_NOT_FOUND, lang=lang)
    
    if request.method == 'GET':
        data = {
            'id': upload.id,
            'academic_year_id': upload.academic_year.id,
            'academic_year_name': upload.academic_year.name,
            'term_id': upload.term.id if upload.term else None,
            'term_name': upload.term.name if upload.term else None,
            'school_level_id': upload.school_level.id,
            'school_level_name': upload.school_level.name,
            'class_level_id': upload.class_level.id,
            'class_level_name': upload.class_level.name,
            'subject_id': upload.subject.id,
            'subject_name': upload.subject.name,
            'grade_type': upload.grade_type,
            'grade_type_display': upload.get_grade_type_display(),
            'weight_percentage': float(upload.weight_percentage),
            'assessment_date': str(upload.assessment_date) if upload.assessment_date else None,
            'excel_file': upload.excel_file.url if upload.excel_file else None,
            'status': upload.status,
            'rejection_reason': upload.rejection_reason,
            'admin_notes': upload.admin_notes,
            'grades_count': upload.student_grades.count(),
            'teacher_name': upload.teacher.full_name if upload.teacher else 'Admin',
            'created_at': upload.created_at.isoformat(),
            'reviewed_at': upload.reviewed_at.isoformat() if upload.reviewed_at else None,
            'reviewed_by_name': upload.reviewed_by.username if upload.reviewed_by else None,
        }
        return _ok(data=data, message="Grade upload details fetched successfully", lang=lang)
    
    elif request.method == 'DELETE':
        if user.role != 'admin':
            return _err("Only admins can delete grade uploads", status.HTTP_403_FORBIDDEN)
        
        try:
            upload_name = f"{upload.subject.name} - {upload.get_grade_type_display()}"
            upload.delete()
            return _ok(message=f"Grade upload '{upload_name}' deleted successfully", lang=lang)
        except Exception as exc:
            return _err(f"Error deleting grade upload: {str(exc)}", status.HTTP_500_INTERNAL_SERVER_ERROR, lang=lang)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_grade_upload_file(request, upload_id):
    """
    Download the original Excel file for a grade upload.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    try:
        upload = GradeUpload.objects.get(
            id=upload_id,
            teacher=teacher
        )
    except GradeUpload.DoesNotExist:
        return _err("Grade upload not found", status.HTTP_404_NOT_FOUND)
    
    if not upload.excel_file or not upload.excel_file.path:
        return _err("File not found", status.HTTP_404_NOT_FOUND)
    
    try:
        with open(upload.excel_file.path, 'rb') as f:
            file_data = f.read()
        
        filename = os.path.basename(upload.excel_file.name)
        response = HttpResponse(
            file_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return _err(f"Error reading file: {str(e)}", status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
    
    
    
# Add these functions to academics_records/views.py

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_grades(request):
    """
    Get all grades for the logged-in teacher's students.
    Used by the teacher grades page.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    # Get query params for pagination
    page = int(request.query_params.get('page', 1))
    page_size = min(int(request.query_params.get('page_size', 10)), 100)
    
    # Get grades from this teacher's uploads
    grades = StudentGrade.objects.filter(
        grade_upload__teacher=teacher
    ).select_related(
        'grade_upload', 'student', 'grade_upload__academic_year',
        'grade_upload__term', 'grade_upload__class_level',
        'grade_upload__subject', 'grade_upload__school_level'
    ).order_by('-created_at')
    
    # Apply filters
    academic_year_id = request.query_params.get('academic_year_id')
    if academic_year_id:
        grades = grades.filter(grade_upload__academic_year_id=academic_year_id)
    
    term_id = request.query_params.get('term_id')
    if term_id:
        grades = grades.filter(grade_upload__term_id=term_id)
    
    class_level_id = request.query_params.get('class_level_id')
    if class_level_id:
        grades = grades.filter(grade_upload__class_level_id=class_level_id)
    
    subject_id = request.query_params.get('subject_id')
    if subject_id:
        grades = grades.filter(grade_upload__subject_id=subject_id)
    
    grade_type = request.query_params.get('grade_type')
    if grade_type:
        grades = grades.filter(grade_upload__grade_type=grade_type)
    
    # Pagination
    total = grades.count()
    start = (page - 1) * page_size
    paginated_grades = grades[start:start + page_size]
    
    # Serialize
    result = []
    for grade in paginated_grades:
        upload = grade.grade_upload
        percentage = (grade.score / grade.max_score * 100) if grade.max_score > 0 else 0
        
        result.append({
            'id': grade.id,
            'student_id': grade.student.id,
            'student_name': grade.student.full_name,
            'student_roll': grade.student.roll_number,
            'academic_year_id': upload.academic_year.id,
            'academic_year_name': upload.academic_year.name,
            'term_id': upload.term.id if upload.term else None,
            'term_name': upload.term.name if upload.term else None,
            'school_level_id': upload.school_level.id,
            'school_level_name': upload.school_level.name,
            'class_level_id': upload.class_level.id,
            'class_level_name': upload.class_level.name,
            'subject_id': upload.subject.id,
            'subject_name': upload.subject.name,
            'grade_type': upload.grade_type,
            'grade_type_display': upload.get_grade_type_display(),
            'score': float(grade.score),
            'max_score': float(grade.max_score),
            'percentage': round(percentage, 2),
            'custom_grade_letter': grade.custom_grade_letter,
            'remarks': grade.remarks,
            'is_published': grade.is_published,
            'published_at': grade.published_at.isoformat() if grade.published_at else None,
            'grade_upload_id': upload.id,
            'upload_status': upload.status,
            'created_at': grade.created_at.isoformat(),
            'updated_at': grade.updated_at.isoformat(),
        })
    
    return Response({
        'success': True,
        'message': t('grades_fetched', lang),
        'data': {
            'results': result,
            'count': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_manual_grade(request):
    from django.db import transaction, IntegrityError
    from decimal import Decimal, InvalidOperation
    from datetime import date
    from django.utils import timezone

    lang = get_lang(request)

    if not _is_admin(request.user):
        return _err(t("forbidden", lang), status.HTTP_403_FORBIDDEN)

    academic_year_id = request.data.get('academic_year_id')
    term_id          = request.data.get('term_id')
    school_level_id  = request.data.get('school_level_id')
    class_level_id   = request.data.get('class_level_id')
    subject_id       = request.data.get('subject_id')
    grade_type       = request.data.get('grade_type')
    student_id       = request.data.get('student_id')
    score            = request.data.get('score')

    missing = [
        name for name, val in [
            ('academic_year_id', academic_year_id),
            ('term_id',          term_id),
            ('school_level_id',  school_level_id),
            ('class_level_id',   class_level_id),
            ('subject_id',       subject_id),
            ('grade_type',       grade_type),
            ('student_id',       student_id),
            ('score',            score),
        ] if not val
    ]
    if missing:
        return _err(
            f"Missing required fields: {', '.join(missing)}",
            status.HTTP_400_BAD_REQUEST
        )

    valid_grade_types = [c[0] for c in GradeType.choices]
    if grade_type not in valid_grade_types:
        return _err(
            f"Invalid grade_type. Must be one of: {', '.join(valid_grade_types)}",
            status.HTTP_400_BAD_REQUEST
        )

    # Validate score before entering the transaction
    try:
        score_decimal = Decimal(str(score))
        if not (Decimal('0') <= score_decimal <= Decimal('100')):
            return _err("Score must be between 0 and 100.", status.HTTP_400_BAD_REQUEST)
    except InvalidOperation:
        return _err("Score must be a valid number.", status.HTTP_400_BAD_REQUEST)

    try:
        max_score = Decimal(str(request.data.get('max_score', 100)))
    except InvalidOperation:
        return _err("max_score must be a valid number.", status.HTTP_400_BAD_REQUEST)

    weight = None
    if request.data.get('weight_percentage'):
        try:
            weight = Decimal(str(request.data.get('weight_percentage')))
        except InvalidOperation:
            return _err("weight_percentage must be a valid number.", status.HTTP_400_BAD_REQUEST)
    if not weight:
        weight = GradeType.get_default_weight(grade_type)

    remarks = request.data.get('remarks', '')

    try:
        academic_year = AcademicYear.objects.get(pk=academic_year_id)
    except AcademicYear.DoesNotExist:
        return _err(f"Academic year not found (id={academic_year_id}).", status.HTTP_404_NOT_FOUND)
    try:
        term = Term.objects.get(pk=term_id)
    except Term.DoesNotExist:
        return _err(f"Term not found (id={term_id}).", status.HTTP_404_NOT_FOUND)
    try:
        school_level = SchoolLevel.objects.get(pk=school_level_id)
    except SchoolLevel.DoesNotExist:
        return _err(f"School level not found (id={school_level_id}).", status.HTTP_404_NOT_FOUND)
    try:
        class_level = ClassLevel.objects.get(pk=class_level_id)
    except ClassLevel.DoesNotExist:
        return _err(f"Class level not found (id={class_level_id}).", status.HTTP_404_NOT_FOUND)
    try:
        subject = Subject.objects.get(pk=subject_id)
    except Subject.DoesNotExist:
        return _err(f"Subject not found (id={subject_id}).", status.HTTP_404_NOT_FOUND)
    try:
        student = Student.objects.get(pk=student_id)
    except Student.DoesNotExist:
        return _err(f"Student not found (id={student_id}).", status.HTTP_404_NOT_FOUND)

    try:
        with transaction.atomic():
            # Find or create the grade upload for this admin/combination
            grade_upload = GradeUpload.objects.filter(
                teacher__isnull=True,
                academic_year=academic_year,
                term=term,
                school_level=school_level,
                class_level=class_level,
                subject=subject,
                grade_type=grade_type,
            ).first()

            if not grade_upload:
                grade_upload = GradeUpload.objects.create(
                    teacher=None,
                    academic_year=academic_year,
                    term=term,
                    school_level=school_level,
                    class_level=class_level,
                    subject=subject,
                    grade_type=grade_type,
                    weight_percentage=weight,
                    assessment_date=date.today(),
                    status=GradeUploadStatus.APPROVED,
                    uploaded_by=request.user,
                )

            existing_grade = StudentGrade.objects.filter(
                grade_upload=grade_upload,
                student=student,
            ).first()

            if existing_grade:
                existing_grade.score               = score_decimal
                existing_grade.max_score           = max_score
                existing_grade.remarks             = remarks
                existing_grade.custom_grade_letter = ''
                existing_grade.is_published        = True
                existing_grade.published_at        = timezone.now()
                existing_grade.save()
                grade   = existing_grade
                message = "Grade updated successfully."
            else:
                grade = StudentGrade.objects.create(
                    grade_upload=grade_upload,
                    student=student,
                    score=score_decimal,
                    max_score=max_score,
                    remarks=remarks,
                    is_published=True,
                    published_at=timezone.now(),
                )
                message = "Grade created successfully."

        # Notify student (outside atomic block so a notification failure
        # doesn't roll back the saved grade)
        if student.user:
            _notify_user(
                student.user,
                'grade_added',
                "New Grade Added",
                f"You received {score_decimal}/{max_score} in {subject.name}",
                created_by=request.user,
                extra_data={
                    'grade_id': grade.id,
                    'subject': subject.name,
                    'score': str(score_decimal),
                },
            )

        percentage = float((score_decimal / max_score) * 100) if max_score > 0 else 0

        return _ok(
            data={
                'grade_id':           grade.id,
                'grade_upload_id':    grade_upload.id,
                'student_id':         student.id,
                'student_name':       student.full_name,
                'student_roll':       student.roll_number,
                'score':              float(score_decimal),
                'max_score':          float(max_score),
                'percentage':         round(percentage, 2),
                'grade_letter':       GradeCalculator.get_grade_letter(Decimal(str(percentage))),
                'remarks':            remarks,
                'subject_name':       subject.name,
                'class_level_name':   class_level.name,
                'academic_year_name': academic_year.name,
                'term_name':          term.name,
                'grade_type':         grade_type,
                'grade_type_display': grade_upload.get_grade_type_display(),
                'weight_percentage':  float(weight),
            },
            message=message,
            status_code=status.HTTP_201_CREATED,
        )

    except IntegrityError as exc:
        print(f"[create_manual_grade] IntegrityError: {exc}")
        return _err(
            "Database integrity error. This usually means a required field is missing "
            "or a duplicate record was detected. Please check your input.",
            status.HTTP_400_BAD_REQUEST,
        )
    except Exception as exc:
        print(f"[create_manual_grade] Unexpected error: {exc}")
        return _err(
            "An unexpected error occurred while saving the grade. Please try again.",
            status.HTTP_500_INTERNAL_SERVER_ERROR,
        )   
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def preview_grade_upload_file(request, upload_id):
    """
    Preview the content of an uploaded Excel file.
    Returns the data in JSON format for display in the modal.
    """
    lang = get_lang(request)
    
    # Permission check
    user = request.user
    if user.role not in ['admin', 'teacher']:
        return _err(t("permission_denied", lang), status.HTTP_403_FORBIDDEN, lang=lang)
    
    try:
        if user.role == 'admin':
            upload = get_object_or_404(GradeUpload, id=upload_id)
        else:
            teacher, err = _get_teacher(request.user, lang)
            if err:
                return err
            upload = get_object_or_404(GradeUpload, id=upload_id, teacher=teacher)
    except GradeUpload.DoesNotExist:
        return _err("Grade upload not found", status.HTTP_404_NOT_FOUND, lang=lang)
    
    if not upload.excel_file or not upload.excel_file.path:
        return _err("File not found", status.HTTP_404_NOT_FOUND, lang=lang)
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(upload.excel_file.path, data_only=True)
        ws = wb.active
        
        # Extract headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # Extract data rows (limit to first 100 rows for performance)
        data_rows = []
        max_rows = min(ws.max_row, 101)  # Header + up to 100 data rows
        
        for row_idx in range(2, max_rows + 1):
            row_data = []
            has_data = False
            for col_idx, cell in enumerate(ws[row_idx]):
                value = cell.value
                if value is not None and str(value).strip():
                    has_data = True
                # Format the value nicely
                if value is None:
                    row_data.append('')
                elif isinstance(value, (int, float)):
                    if isinstance(value, float) and value.is_integer():
                        row_data.append(str(int(value)))
                    else:
                        row_data.append(str(value))
                else:
                    row_data.append(str(value).strip())
            
            if has_data:
                data_rows.append(row_data)
        
        # Get file info
        file_size = os.path.getsize(upload.excel_file.path)
        file_name = os.path.basename(upload.excel_file.name)
        
        return Response({
            'success': True,
            'data': {
                'file_name': file_name,
                'file_size': file_size,
                'file_size_mb': round(file_size / (1024 * 1024), 2),
                'total_rows': len(data_rows),
                'headers': headers,
                'data_rows': data_rows,
                'has_more': ws.max_row > max_rows
            }
        })
        
    except Exception as exc:
        return _err(f"Error reading file: {str(exc)}", status.HTTP_500_INTERNAL_SERVER_ERROR, lang=lang)
    
    
    
    
    
    
# ============================================================
# ATTENDANCE SESSION VIEWS (UPDATED)
# ============================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_attendance_sessions(request):
    """
    Get all attendance sessions for the logged-in teacher with filtering.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    # Base queryset
    sessions = AttendanceSession.objects.filter(
        teacher=teacher,
        is_submitted=True
    ).select_related(
        'academic_year', 'term', 'school_level', 'class_level',
        'subject', 'classroom', 'teacher'
    ).order_by('-session_date')
    
    # Apply filters
    academic_year_id = request.query_params.get('academic_year_id')
    if academic_year_id:
        sessions = sessions.filter(academic_year_id=academic_year_id)
    
    term_id = request.query_params.get('term_id')
    if term_id:
        sessions = sessions.filter(term_id=term_id)
    
    class_level_id = request.query_params.get('class_level_id')
    if class_level_id:
        sessions = sessions.filter(class_level_id=class_level_id)
    
    subject_id = request.query_params.get('subject_id')
    if subject_id:
        sessions = sessions.filter(subject_id=subject_id)
    
    classroom_id = request.query_params.get('classroom_id')
    if classroom_id:
        sessions = sessions.filter(classroom_id=classroom_id)
    
    # Serialize
    from .serializers import AttendanceSessionListSerializer
    serializer = AttendanceSessionListSerializer(sessions, many=True)
    
    return Response({
        'success': True,
        'message': t('attendance_sessions_fetched', lang),
        'data': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([JSONParser])
def create_attendance_session(request):
    """
    Create a new attendance session with records.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    # Validate required fields
    academic_year_id = request.data.get('academic_year_id')
    class_level_id = request.data.get('class_level_id')
    subject_id = request.data.get('subject_id')
    session_date = request.data.get('session_date')
    records = request.data.get('records', [])
    
    if not all([academic_year_id, class_level_id, subject_id, session_date]):
        return Response({
            'success': False,
            'message': 'Missing required fields: academic_year_id, class_level_id, subject_id, session_date'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not records:
        return Response({
            'success': False,
            'message': 'No attendance records to submit'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        subject = get_object_or_404(Subject, id=subject_id)
        school_level = class_level.school_level
        
        term_id = request.data.get('term_id')
        term = get_object_or_404(Term, id=term_id) if term_id else None
        
        classroom_id = request.data.get('classroom_id')
        classroom = get_object_or_404(ClassRoom, id=classroom_id) if classroom_id else None
        
        # Check for duplicate session
        if AttendanceSession.objects.filter(
            teacher=teacher,
            academic_year=academic_year,
            class_level=class_level,
            subject=subject,
            session_date=session_date
        ).exists():
            return Response({
                'success': False,
                'message': 'Attendance already recorded for this session'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create session
        with transaction.atomic():
            session = AttendanceSession.objects.create(
                teacher=teacher,
                academic_year=academic_year,
                term=term,
                school_level=school_level,
                class_level=class_level,
                subject=subject,
                classroom=classroom,
                session_date=session_date,
                start_time=request.data.get('start_time'),
                end_time=request.data.get('end_time'),
                notes=request.data.get('notes', ''),
                is_submitted=True,
                submitted_at=timezone.now(),
                created_by=request.user
            )
            
            # Create attendance records
            for record in records:
                student_id = record.get('student_id')
                status_value = record.get('status', 'present')
                remarks = record.get('remarks', '')
                
                student = get_object_or_404(Student, id=student_id)
                
                StudentAttendance.objects.create(
                    session=session,
                    student=student,
                    status=status_value,
                    remarks=remarks
                )
        
        return Response({
            'success': True,
            'message': 'Attendance recorded successfully',
            'data': {
                'session_id': session.id,
                'records_count': len(records)
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as exc:
        return Response({
            'success': False,
            'message': f'An unexpected error occurred: {str(exc)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# academics_records/views.py

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attendance_session_detail(request, session_id):
    """
    Get detailed attendance session with all records.
    """
    import json
    lang = get_lang(request)
    
    print(f"\n{'='*60}")
    print(f"📥 ATTENDANCE SESSION DETAIL REQUEST")
    print(f"   Session ID: {session_id}")
    print(f"   User: {request.user.username} (role={request.user.role})")
    print(f"{'='*60}")
    
    # Check if user is admin or teacher
    if _is_admin(request.user):
        print(f"   👑 Admin access - can view any session")
        try:
            session = AttendanceSession.objects.get(id=session_id)
        except AttendanceSession.DoesNotExist:
            print(f"   ❌ Session {session_id} not found")
            return Response({
                'success': False,
                'message': 'Attendance session not found'
            }, status=status.HTTP_404_NOT_FOUND)
    else:
        teacher, err = _get_teacher(request.user, lang)
        if err:
            return err
        
        print(f"   👨‍🏫 Teacher access - teacher_id={teacher.id}")
        
        try:
            session = AttendanceSession.objects.get(
                id=session_id,
                teacher=teacher
            )
        except AttendanceSession.DoesNotExist:
            print(f"   ❌ Session {session_id} not found for this teacher")
            return Response({
                'success': False,
                'message': 'Attendance session not found'
            }, status=status.HTTP_404_NOT_FOUND)
    
    print(f"   ✅ Session found:")
    print(f"      - Subject: {session.subject.name}")
    print(f"      - Class: {session.class_level.name}")
    print(f"      - Date: {session.session_date}")
    print(f"      - Records count: {session.records.count()}")
    
    # Serialize the session with its records
    from .serializers import AttendanceSessionDetailSerializer
    serializer = AttendanceSessionDetailSerializer(session)
    
    print(f"   📤 Serialized data keys: {list(serializer.data.keys())}")
    print(f"   📅 session_date in response: {serializer.data.get('session_date')}")
    print(f"{'='*60}\n")
    
    return Response({
        'success': True,
        'data': serializer.data
    })

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_attendance_record(request, record_id):
    """
    Update a single attendance record.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    try:
        record = StudentAttendance.objects.get(
            id=record_id,
            session__teacher=teacher
        )
    except StudentAttendance.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Attendance record not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    status_value = request.data.get('status')
    remarks = request.data.get('remarks', record.remarks)
    
    if status_value:
        record.status = status_value
    record.remarks = remarks
    record.save()
    
    return Response({
        'success': True,
        'message': 'Attendance record updated successfully',
        'data': {
            'id': record.id,
            'status': record.status,
            'remarks': record.remarks
        }
    })


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_attendance_session(request, session_id):
    """
    Delete an attendance session and all its records.
    """
    lang = get_lang(request)
    
    # Check if user is admin
    if _is_admin(request.user):
        try:
            session = AttendanceSession.objects.get(id=session_id)
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Attendance session not found'
            }, status=status.HTTP_404_NOT_FOUND)
    else:
        teacher, err = _get_teacher(request.user, lang)
        if err:
            return err
        
        try:
            session = AttendanceSession.objects.get(
                id=session_id,
                teacher=teacher
            )
        except AttendanceSession.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Attendance session not found'
            }, status=status.HTTP_404_NOT_FOUND)
    
    # Store session info for response message
    session_info = f"{session.subject.name} - {session.session_date}"
    session.delete()
    
    return Response({
        'success': True,
        'message': f'Attendance session for {session_info} deleted successfully'
    })
    
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_teacher_students_for_classroom(request, classroom_id):
    """
    Get all students in a specific classroom for the logged-in teacher.
    """
    lang = get_lang(request)
    
    teacher, err = _get_teacher(request.user, lang)
    if err:
        return err
    
    try:
        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        
        # Verify teacher is assigned to this classroom
        is_assigned = TeacherAssignment.objects.filter(
            teacher=teacher,
            classrooms=classroom,
            status='active'
        ).exists()
        
        if not is_assigned:
            return Response({
                'success': False,
                'message': 'You are not assigned to this classroom'
            }, status=status.HTTP_403_FORBIDDEN)
        
        academic_year_id = request.query_params.get('academic_year_id')
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id) if academic_year_id else None
        
        if not academic_year:
            academic_year = AcademicYear.objects.filter(is_current=True).first()
        
        # Get students in this classroom for the current academic year
        from students.models import StudentClassroomAssignment
        
        students = Student.objects.filter(
            studentclassroomassignment__classroom=classroom,
            studentclassroomassignment__academic_year=academic_year,
            studentclassroomassignment__status='active',
            status='active'
        ).distinct().order_by('full_name')
        
        # Serialize
        result = []
        for student in students:
            result.append({
                'id': student.id,
                'full_name': student.full_name,
                'roll_number': student.roll_number,
                'email': student.email,
                'phone_number': student.phone_number
            })
        
        return Response({
            'success': True,
            'data': {
                'classroom_id': classroom.id,
                'classroom_name': classroom.name,
                'academic_year': academic_year.name if academic_year else None,
                'students': result,
                'total': len(result)
            }
        })
        
    except ClassRoom.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Classroom not found'
        }, status=status.HTTP_404_NOT_FOUND)
        
        
        
        
        
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_full_report(request, student_id):
    """
    Get full academic report for a student across all terms in the academic year.
    """
    lang = get_lang(request)
    
    # Check permissions
    user = request.user
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        return _err("Student not found", status.HTTP_404_NOT_FOUND)
    
    if user.role == 'parent':
        parent = get_object_or_404(Parent, user=user)
        if not parent.students.filter(id=student_id).exists():
            return _err(t("permission_denied", lang), status.HTTP_403_FORBIDDEN)
    elif user.role != 'admin' and user.role != 'student':
        return _err(t("permission_denied", lang), status.HTTP_403_FORBIDDEN)
    
    academic_year_id = request.query_params.get('academic_year_id')
    if not academic_year_id:
        academic_year = AcademicYear.objects.filter(is_current=True).first()
    else:
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
    
    # Get all terms for this academic year
    terms = Term.objects.filter(academic_year=academic_year).order_by('term_number')
    
    term_performances = []
    for term in terms:
        performance = PerformanceReportGenerator.get_full_report(student, academic_year, term)
        term_performances.append({
            'term_id': term.id,
            'term_name': term.name,
            'term_number': term.term_number,
            'is_current': term.is_current,
            'overall_average': performance['academic_performance']['overall_average'],
            'grade_letter': performance['academic_performance']['grade_letter'],
            'subjects_passed': performance['academic_performance']['subjects_passed'],
            'subjects_failed': performance['academic_performance']['subjects_failed'],
            'total_subjects': performance['academic_performance']['total_subjects'],
            'subject_results': performance['academic_performance']['subject_results'],
            'discipline': performance['discipline']
        })
    
    # Get current term performance
    current_term = terms.filter(is_current=True).first()
    current_performance = next((t for t in term_performances if t['is_current']), term_performances[0] if term_performances else None)
    
    return Response({
        'success': True,
        'data': {
            'student_id': student.id,
            'student_name': student.full_name,
            'student_roll': student.roll_number,
            'academic_year_id': academic_year.id,
            'academic_year_name': academic_year.name,
            'current_term_performance': current_performance,
            'term_performances': term_performances,
            'generated_at': timezone.now().isoformat()
        }
    })
    
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_full_report(request):
    """
    Get full academic report for the currently logged-in student across all terms in the academic year.
    """
    lang = get_lang(request)
    
    # Check if the logged-in user is a student
    user = request.user
    
    if user.role != 'student':
        return _err(t("permission_denied", lang), status.HTTP_403_FORBIDDEN)
    
    try:
        student = Student.objects.get(user=user)
    except Student.DoesNotExist:
        return _err("Student profile not found", status.HTTP_404_NOT_FOUND)
    
    academic_year_id = request.query_params.get('academic_year_id')
    if not academic_year_id:
        academic_year = AcademicYear.objects.filter(is_current=True).first()
        if not academic_year:
            return _err("No current academic year found", status.HTTP_404_NOT_FOUND)
    else:
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
    
    # Get all terms for this academic year
    terms = Term.objects.filter(academic_year=academic_year).order_by('term_number')
    
    if not terms.exists():
        return _err("No terms found for this academic year", status.HTTP_404_NOT_FOUND)
    
    term_performances = []
    for term in terms:
        performance = PerformanceReportGenerator.get_full_report(student, academic_year, term)
        term_performances.append({
            'term_id': term.id,
            'term_name': term.name,
            'term_number': term.term_number,
            'is_current': term.is_current,
            'overall_average': performance['academic_performance']['overall_average'],
            'grade_letter': performance['academic_performance']['grade_letter'],
            'subjects_passed': performance['academic_performance']['subjects_passed'],
            'subjects_failed': performance['academic_performance']['subjects_failed'],
            'total_subjects': performance['academic_performance']['total_subjects'],
            'subject_results': performance['academic_performance']['subject_results'],
            'discipline': performance['discipline']
        })
    
    # Get current term performance
    current_performance = next((t for t in term_performances if t['is_current']), term_performances[0] if term_performances else None)
    
    return Response({
        'success': True,
        'data': {
            'student_id': student.id,
            'student_name': student.full_name,
            'student_roll': student.roll_number,
            'academic_year_id': academic_year.id,
            'academic_year_name': academic_year.name,
            'current_term_performance': current_performance,
            'term_performances': term_performances,
            'generated_at': timezone.now().isoformat()
        }
    })